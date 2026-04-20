from __future__ import annotations

import json
import queue
import re
import shutil
import threading
import zipfile
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from .config import AppConfig
from .oee_state import build_work_order_snapshot
from .parsers import (
    normalize_color,
    normalize_token,
    parse_colon_value_line,
    parse_mega_event_from_log,
    parse_tablet_fault_line,
    parse_tablet_oee_line,
    parse_vision_event,
)


EVENT_LOG_COLUMNS = [
    "log_event_id", "event_time_text", "source_id", "source_code", "event_type_id", "event_type_code", "event_summary_tr",
    "item_id", "measure_id", "fault_id", "oee_snapshot_id", "vision_event_id", "line_id", "station_id", "color_id",
    "color_code", "decision_source_id", "decision_source_code", "mega_state_id", "mega_state_code", "queue_depth",
    "review_required", "travel_ms", "notes", "raw_line",
]
MEASUREMENT_COLUMNS = [
    "measurement_row_id", "measure_id", "item_id", "measured_at", "measurement_log_event_id", "source_log_file",
    "final_color_id", "final_color_code", "final_color_raw", "decision_source_id", "decision_source_code",
    "search_hint", "search_hint_win", "search_hint_second", "search_hint_strong", "search_hint_fallback_allowed",
    "review_required", "core_used", "core_n", "obj_n", "median_nearest", "score_nearest", "med_r", "med_g", "med_b",
    "med_d_r", "med_d_y", "med_d_b", "med_d_x", "x_r", "x_g", "x_b", "med_obj", "confidence", "core_str_min",
    "core_str_max", "vote_win", "vote_second", "vote_classified", "vote_x", "vote_r", "vote_y", "vote_b", "vote_cal",
    "tot_r", "tot_y", "tot_b", "tot_x", "tot_cal", "measurement_error_flag", "measurement_error_reason", "raw_line",
]
COMPLETED_COLUMNS = [
    "production_record_id", "item_id", "measure_id", "queue_event_log_id", "completion_event_log_id", "detected_at",
    "completed_at", "color_id", "color_code", "color_raw", "status_code", "status_tr", "travel_ms", "flow_ms", "cycle_ms",
    "decision_source_id", "decision_source_code", "review_required", "final_quality_code", "final_quality_tr",
    "override_flag", "override_source_code", "override_applied_at", "sensor_color_code", "vision_color_code",
    "final_color_code", "mismatch_flag", "correlation_status", "finalization_reason", "early_pick_triggered",
    "pick_trigger_source", "early_pick_request_sent_at", "early_pick_accepted_at", "final_color_frozen_at",
]
OEE_SNAPSHOT_COLUMNS = [
    "oee_snapshot_id", "snapshot_time", "event_log_id", "sample_cycle_tag", "oee", "availability", "performance", "quality",
    "mavi_s", "mavi_r", "mavi_h", "sari_s", "sari_r", "sari_h", "kirmizi_s", "kirmizi_r", "kirmizi_h",
    "is_full_cycle_reference", "notes", "raw_line",
]
VISION_COLUMNS = [
    "vision_event_id", "event_time", "source_code", "vision_track_id", "event_type", "color_id", "color_code", "item_id",
    "measure_id", "confidence", "confidence_tier", "correlation_status", "late_vision_audit_flag", "vision_observed_at",
    "vision_published_at", "vision_received_at", "line_id", "station_id", "bbox_x1", "bbox_y1", "bbox_x2", "bbox_y2",
    "direction", "is_placeholder", "notes",
]
RAW_LOG_COLUMNS = [
    "raw_log_id", "logged_at", "source_topic", "source_code", "parsed_flag", "event_type_code", "item_id", "measure_id",
    "color_code", "notes", "raw_payload",
]
WORK_ORDER_COLUMNS = [
    "work_order_record_id", "order_id", "erp_type", "source_file", "queued_at", "started_at", "completed_at",
    "status_code", "status_tr", "stock_code", "stock_name", "product_color", "target_qty", "fulfilled_qty",
    "production_qty", "inventory_consumed_qty", "good_qty", "rework_qty", "scrap_qty", "ideal_cycle_ms",
    "ideal_cycle_sec", "planned_duration_ms", "planned_duration_sec", "runtime_ms", "runtime_sec",
    "unplanned_downtime_ms", "unplanned_downtime_sec", "availability_pct", "performance_pct", "quality_pct",
    "oee_pct", "started_by", "started_by_name", "transition_reason", "notes",
]
INVENTORY_COLUMNS = [
    "inventory_record_id", "match_key", "product_code", "stock_code", "stock_name", "color_code", "quantity",
    "last_updated_at", "last_source", "source_file", "notes",
]
FAULT_COLUMNS = [
    "fault_record_id", "fault_id", "status_code", "status_tr", "fault_type_code", "fault_category", "fault_reason_tr",
    "started_at", "ended_at", "duration_ms", "duration_sec", "source_code", "device_id", "device_name",
    "bound_station_id", "operator_id", "operator_code", "operator_name", "shift_code", "notes",
]
MAINTENANCE_COLUMNS = [
    "maintenance_record_id", "maintenance_row_key", "session_id", "phase_code", "phase_tr", "step_code", "step_label_tr",
    "required_flag", "completed_flag", "completed_at", "session_started_at", "session_ended_at", "duration_ms",
    "duration_sec", "shift_code", "device_id", "device_name", "device_role", "bound_station_id", "operator_id",
    "operator_code", "operator_name", "note",
]

RAW_LOG_SHEET_NAME = "99_Raw_Logs"
LEGACY_RAW_LOG_SHEET_NAME = "7_Raw_Logs"
WORK_ORDER_SHEET_NAME = "7_Is_Emirleri"
LEGACY_WORK_ORDER_SHEET_NAME = "8_Is_Emirleri"
INVENTORY_SHEET_NAME = "8_Depo_Stok"
LEGACY_INVENTORY_SHEET_NAME = "9_Depo_Stok"
MAINTENANCE_SHEET_NAME = "9_Bakim_Kayitlari"

SHEET_COLUMNS = {
    "1_Olay_Logu": EVENT_LOG_COLUMNS,
    "2_Olcumler": MEASUREMENT_COLUMNS,
    "3_Arizalar": FAULT_COLUMNS,
    "4_Uretim_Tamamlanan": COMPLETED_COLUMNS,
    "5_OEE_Anliklari": OEE_SNAPSHOT_COLUMNS,
    "6_Vision": VISION_COLUMNS,
    RAW_LOG_SHEET_NAME: RAW_LOG_COLUMNS,
    WORK_ORDER_SHEET_NAME: WORK_ORDER_COLUMNS,
    INVENTORY_SHEET_NAME: INVENTORY_COLUMNS,
    MAINTENANCE_SHEET_NAME: MAINTENANCE_COLUMNS,
}

SOURCE_IDS = {"mega": 1, "tablet": 2, "vision": 3, "system": 4}
COLOR_IDS = {"red": 1, "yellow": 2, "blue": 3, "empty": 4, "uncertain": 5}
EVENT_TYPE_IDS = {
    "measurement_decision": 2,
    "queue_enq": 3,
    "arm_position_reached": 4,
    "pickplace_done": 5,
    "pick_command_rejected": 6,
    "pick_released": 7,
    "pick_return_started": 8,
    "pick_return_reached": 9,
    "pickplace_return_done": 10,
    "vision_event": 11,
    "early_pick_request": 12,
    "tablet_oee_snapshot": 13,
    "tablet_fault": 14,
    "oee_control": 15,
    "shift_start": 16,
    "shift_stop": 17,
    "counts_reset": 18,
    "pickplace_started": 19,
    "pick_drop_reached": 20,
    "maintenance_opening_started": 21,
    "maintenance_step_completed": 22,
    "maintenance_completed": 23,
    "maintenance_closing_started": 24,
    "kiosk_fault_started": 25,
    "kiosk_fault_cleared": 26,
    "help_requested": 27,
    "help_acknowledged": 28,
    "help_resolved": 29,
}
DECISION_SOURCE_IDS = {"CORE_STABLE": 1, "MEDIAN_STABLE": 2, "CORE_VOTE_MATCH": 3, "VISION": 4, "TABLET": 5, "SYSTEM": 6}
MEGA_STATE_IDS = {"SEARCH": 1, "SEARCHING": 2, "MEASURING": 3, "WAIT_ARM": 4, "PAUSED": 5, "STOPPED": 6, "QUEUE": 7}
EXCEL_ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def _station_for_event(event_type_code: str) -> int:
    if event_type_code in {"measurement_decision"}:
        return 1
    if event_type_code == "queue_enq":
        return 2
    if event_type_code in {"arm_position_reached", "pickplace_started", "pick_drop_reached", "pickplace_done", "pick_command_rejected", "pick_released", "pick_return_started", "pick_return_reached", "pickplace_return_done", "early_pick_request"}:
        return 3
    if event_type_code in {
        "maintenance_opening_started",
        "maintenance_step_completed",
        "maintenance_completed",
        "maintenance_closing_started",
        "kiosk_fault_started",
        "kiosk_fault_cleared",
        "help_requested",
        "help_acknowledged",
        "help_resolved",
    }:
        return 4
    return 6 if event_type_code == "vision_event" else 5


def _decision_source_id(value: str | None) -> int | str:
    code = str(value or "").strip().upper()
    return DECISION_SOURCE_IDS.get(code, "")


def _color_id(value: str | None) -> int | str:
    return COLOR_IDS.get(str(value or "").strip().lower(), "")


def _mega_state_id(value: str | None) -> int | str:
    code = str(value or "").strip().upper()
    return MEGA_STATE_IDS.get(code, "")


def _json_text(payload: Any) -> str:
    if isinstance(payload, str):
        return payload
    return json.dumps(payload, ensure_ascii=False, sort_keys=True)


def _excel_cell_value(value: Any) -> Any:
    if isinstance(value, str):
        return EXCEL_ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _safe_int(value: Any) -> int | str:
    if value in (None, ""):
        return ""
    try:
        return int(value)
    except (TypeError, ValueError):
        return ""


def _optional_pct(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return round(float(value) * 100.0, 1)
    except (TypeError, ValueError):
        return None


def _parse_iso_text(value: Any) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None


def _duration_seconds(start_at: Any, end_at: Any) -> float | None:
    duration_ms = _duration_milliseconds(start_at, end_at)
    if duration_ms is None:
        return None
    return round(duration_ms / 1000.0, 1)


def _duration_milliseconds(start_at: Any, end_at: Any) -> int | None:
    start_dt = _parse_iso_text(start_at)
    end_dt = _parse_iso_text(end_at)
    if start_dt is None or end_dt is None or end_dt < start_dt:
        return None
    return int((end_dt - start_dt).total_seconds() * 1000)


def _measurement_error_info(final_color: str, confidence: Any) -> tuple[int, str]:
    if final_color in {"empty", "uncertain"}:
        return 1, "final_color_invalid"
    text = str(confidence or "").strip()
    if text in {"0", "0.0", "false", "False"}:
        return 1, "confidence=0"
    return 0, ""


def _color_label_tr(color_code: Any) -> str:
    return {
        "red": "kirmizi",
        "yellow": "sari",
        "blue": "mavi",
        "empty": "bos",
        "uncertain": "belirsiz",
    }.get(str(color_code or "").strip().lower(), str(color_code or "").strip().lower() or "-")


def _vision_event_label_tr(event_type: Any) -> str:
    return {
        "line_crossed": "hat gecisi",
        "box_confirmed": "kutu dogrulama",
        "track_started": "iz basladi",
        "track_lost": "iz kayboldu",
    }.get(str(event_type or "").strip().lower(), str(event_type or "").strip().replace("_", " ") or "vision")


def _correlation_status_label_tr(status_code: Any) -> str:
    return {
        "matched": "eslesti",
        "late": "gec geldi",
        "ambiguous": "belirsiz",
        "drifted": "kuyruk kaydi",
        "ignored_low_conf": "dusuk guvenle yok sayildi",
        "fault_active": "aktif ariza sirasinda yok sayildi",
    }.get(str(status_code or "").strip().lower(), str(status_code or "").strip().replace("_", " "))


def _work_order_status_tr(status_code: Any) -> str:
    return {
        "queued": "Sirada",
        "active": "Aktif",
        "pending_approval": "Onay bekliyor",
        "completed": "Tamamlandi",
        "rolled_back": "Geri alindi",
    }.get(str(status_code or "").strip().lower(), str(status_code or "").strip().upper())


def _fault_status_label_tr(status_code: Any) -> str:
    return {
        "basladi": "basladi",
        "bitti": "bitti",
        "aktif": "aktif",
        "yok": "yok",
    }.get(normalize_token(status_code), str(status_code or "").strip().lower() or "-")


@dataclass(slots=True)
class SinkEnvelope:
    kind: str
    received_at: str
    payload: Any


class WorkbookProjector:
    def __init__(self) -> None:
        self.counters = {
            "log_event_id": 1,
            "measurement_row_id": 1,
            "production_record_id": 1,
            "oee_snapshot_id": 1,
            "vision_event_id": 1,
            "raw_log_id": 1,
        }
        self.completed_state: dict[str, dict[str, Any]] = {}
        self.completed_rows_by_item: dict[str, dict[str, Any]] = {}
        self.last_completion_at = ""
        self.pending_completed_row_update: dict[str, Any] | None = None

    def prime(self, counters: dict[str, int]) -> None:
        self.counters.update(counters)

    def _next(self, key: str) -> int:
        value = self.counters[key]
        self.counters[key] += 1
        return value

    def _completed_key(self, item_id: str | None, measure_id: str | None) -> str:
        normalized_item_id = str(item_id or "").strip()
        if normalized_item_id:
            return normalized_item_id
        normalized_measure_id = str(measure_id or "").strip()
        return f"measure:{normalized_measure_id}" if normalized_measure_id else ""

    def _active_pick_key(self) -> str:
        candidates: list[tuple[str, str]] = []
        for key, state in self.completed_state.items():
            picked_at = str(state.get("picked_at") or "").strip()
            if not picked_at or state.get("released_at"):
                continue
            candidates.append((picked_at, key))
        if not candidates:
            return ""
        candidates.sort()
        return candidates[-1][1]

    def _resolve_robot_event_context(self, parsed: dict[str, Any]) -> tuple[dict[str, Any], str]:
        key = self._completed_key(parsed["item_id"], parsed["measure_id"])
        if key:
            return parsed, key
        if parsed["event_type"] not in {"pickplace_started", "pick_drop_reached"}:
            return parsed, key
        fallback_key = self._active_pick_key()
        if not fallback_key:
            return parsed, key
        state = self.completed_state.get(fallback_key, {})
        resolved = dict(parsed)
        resolved["item_id"] = str(state.get("item_id") or "")
        resolved["measure_id"] = str(state.get("measure_id") or "")
        resolved["color"] = str(state.get("final_color") or state.get("sensor_color") or parsed.get("color") or "")
        resolved["decision_source"] = str(state.get("sensor_decision_source") or parsed.get("decision_source") or "")
        resolved["review_required"] = bool(state.get("review_required")) or bool(parsed.get("review_required"))
        return resolved, fallback_key

    def _current_event_review_required(self, key: str, parsed: dict[str, Any]) -> bool:
        state = self.completed_state.get(key, {}) if key else {}
        return bool(parsed.get("review_required")) or bool(state.get("review_required"))

    def _current_event_decision_source(self, key: str, parsed: dict[str, Any]) -> str:
        state = self.completed_state.get(key, {}) if key else {}
        return str(state.get("decision_source") or state.get("sensor_decision_source") or parsed.get("decision_source") or "")

    def _current_event_color(self, key: str, parsed: dict[str, Any]) -> str:
        state = self.completed_state.get(key, {}) if key else {}
        return str(state.get("final_color") or parsed.get("color") or "")

    def _find_completed_row(self, item_id: str | None, measure_id: str | None) -> dict[str, Any] | None:
        normalized_item_id = str(item_id or "").strip()
        if normalized_item_id:
            row = self.completed_rows_by_item.get(normalized_item_id)
            if isinstance(row, dict):
                return row
        normalized_measure_id = str(measure_id or "").strip()
        if not normalized_measure_id:
            return None
        for row in self.completed_rows_by_item.values():
            if str(row.get("measure_id") or "").strip() == normalized_measure_id:
                return row
        return None

    def _set_completed_row_vision_fields(self, row: dict[str, Any], parsed: dict[str, Any], received_at: str) -> None:
        row["vision_color_code"] = parsed.get("color") or row.get("vision_color_code", "")
        row["correlation_status"] = parsed.get("correlation_status") or row.get("correlation_status", "")
        if parsed.get("late_vision_audit_flag"):
            row["finalization_reason"] = "SENSOR_LATE_VISION"
        if bool(parsed.get("review_required")):
            row["review_required"] = 1
            if str(row.get("status_code") or "").strip().upper() == "COMPLETED":
                row["status_code"] = "COMPLETED_REVIEW"
                row["status_tr"] = "Inceleme gerekli"
        sensor_color = str(row.get("sensor_color_code") or "").strip().lower()
        vision_color = str(parsed.get("color") or "").strip().lower()
        if sensor_color and vision_color and sensor_color != vision_color:
            row["mismatch_flag"] = 1
        if parsed.get("decision_applied"):
            row["decision_source_code"] = "VISION"
            row["decision_source_id"] = _decision_source_id("VISION")
            row["final_color_code"] = parsed.get("color") or row.get("final_color_code", "")
            row["color_code"] = parsed.get("color") or row.get("color_code", "")
            row["color_raw"] = parsed.get("color") or row.get("color_raw", "")
            row["finalization_reason"] = "VISION_CORRECTED_MISMATCH" if row.get("mismatch_flag") else "VISION_HIGH_CONF"
            row["final_color_frozen_at"] = received_at or row.get("final_color_frozen_at", "")

    def _oee_snapshot_row(self, *, snapshot_id: int, received_at: str, event_log_id: int, parsed_oee: dict[str, Any], raw_line: str, notes: str) -> dict[str, Any]:
        colors = parsed_oee.get("colors") if isinstance(parsed_oee.get("colors"), dict) else {}
        return {
            "oee_snapshot_id": snapshot_id,
            "snapshot_time": received_at,
            "event_log_id": event_log_id,
            "sample_cycle_tag": f"tablet:{snapshot_id}",
            "oee": parsed_oee.get("oee"),
            "availability": parsed_oee.get("availability"),
            "performance": parsed_oee.get("performance"),
            "quality": parsed_oee.get("quality"),
            "mavi_s": int(((colors.get("blue") or {}).get("good") or 0)),
            "mavi_r": int(((colors.get("blue") or {}).get("rework") or 0)),
            "mavi_h": int(((colors.get("blue") or {}).get("scrap") or 0)),
            "sari_s": int(((colors.get("yellow") or {}).get("good") or 0)),
            "sari_r": int(((colors.get("yellow") or {}).get("rework") or 0)),
            "sari_h": int(((colors.get("yellow") or {}).get("scrap") or 0)),
            "kirmizi_s": int(((colors.get("red") or {}).get("good") or 0)),
            "kirmizi_r": int(((colors.get("red") or {}).get("rework") or 0)),
            "kirmizi_h": int(((colors.get("red") or {}).get("scrap") or 0)),
            "is_full_cycle_reference": 1,
            "notes": notes,
            "raw_line": raw_line,
        }

    def _duration_ms(self, start_at: Any, end_at: Any) -> int | str:
        start_text = str(start_at or "").strip()
        end_text = str(end_at or "").strip()
        if not start_text or not end_text:
            return ""
        try:
            from datetime import datetime

            start_dt = datetime.fromisoformat(start_text.replace("Z", "+00:00"))
            end_dt = datetime.fromisoformat(end_text.replace("Z", "+00:00"))
        except ValueError:
            return ""
        delta_ms = int((end_dt - start_dt).total_seconds() * 1000)
        return delta_ms if delta_ms >= 0 else ""

    def _finalize_completed_row(self, *, state: dict[str, Any], parsed: dict[str, Any], key: str, received_at: str, log_event_id: int) -> dict[str, Any]:
        completed_at = str(state.get("released_at") or received_at or "").strip()
        detected_at = str(state.get("detected_at") or state.get("measured_at") or "").strip()
        flow_ms = self._duration_ms(detected_at, completed_at)
        cycle_ms = self._duration_ms(self.last_completion_at, completed_at) if self.last_completion_at else ""
        if completed_at:
            self.last_completion_at = completed_at
        row = {
            "production_record_id": self._next("production_record_id"),
            "item_id": state.get("item_id", parsed["item_id"]),
            "measure_id": state.get("measure_id", parsed["measure_id"]),
            "queue_event_log_id": state.get("queue_event_log_id", ""),
            "completion_event_log_id": log_event_id,
            "detected_at": detected_at,
            "completed_at": completed_at,
            "color_id": _color_id(state.get("final_color", state.get("sensor_color", parsed["color"]))),
            "color_code": state.get("final_color", state.get("sensor_color", parsed["color"])),
            "color_raw": state.get("final_color", state.get("sensor_color", parsed["color"])),
            "status_code": "COMPLETED_REVIEW" if state.get("review_required", parsed["review_required"]) else "COMPLETED",
            "status_tr": "Inceleme gerekli" if state.get("review_required", parsed["review_required"]) else "Tamamlandi",
            "travel_ms": _safe_int(state.get("travel_ms", parsed["travel_ms"])),
            "flow_ms": flow_ms,
            "cycle_ms": cycle_ms,
            "decision_source_id": _decision_source_id(state.get("decision_source", parsed["decision_source"])),
            "decision_source_code": str(state.get("decision_source", parsed["decision_source"]) or "").upper(),
            "review_required": 1 if state.get("review_required", parsed["review_required"]) else 0,
            "final_quality_code": "GOOD",
            "final_quality_tr": "Saglam",
            "override_flag": 0,
            "override_source_code": "",
            "override_applied_at": "",
            "sensor_color_code": state.get("sensor_color", parsed["color"]),
            "vision_color_code": state.get("vision_color", ""),
            "final_color_code": state.get("final_color", state.get("sensor_color", parsed["color"])),
            "mismatch_flag": 1 if state.get("mismatch_flag") else 0,
            "correlation_status": state.get("correlation_status", ""),
            "finalization_reason": state.get("finalization_reason", "SENSOR_NO_VISION"),
            "early_pick_triggered": 1 if state.get("early_pick_triggered") else 0,
            "pick_trigger_source": state.get("pick_trigger_source", str(parsed.get("trigger_source") or "").upper()),
            "early_pick_request_sent_at": state.get("early_pick_request_sent_at", ""),
            "early_pick_accepted_at": state.get("early_pick_accepted_at", ""),
            "final_color_frozen_at": completed_at or received_at,
        }
        self.completed_rows_by_item[str(state.get("item_id", parsed["item_id"]) or key)] = row
        self.completed_state.pop(key, None)
        return row

    def _raw_row(self, *, received_at: str, source_topic: str, source_code: str, raw_payload: Any, parsed_flag: int, event_type_code: str = "", item_id: str = "", measure_id: str = "", color_code: str = "", notes: str = "") -> dict[str, Any]:
        return {
            "raw_log_id": self._next("raw_log_id"),
            "logged_at": received_at,
            "source_topic": source_topic,
            "source_code": source_code,
            "parsed_flag": parsed_flag,
            "event_type_code": event_type_code,
            "item_id": item_id,
            "measure_id": measure_id,
            "color_code": color_code,
            "notes": notes,
            "raw_payload": _json_text(raw_payload),
        }

    def _event_row(self, *, log_event_id: int, received_at: str, source_code: str, event_type_code: str, item_id: str = "", measure_id: str = "", color_code: str = "", decision_source_code: str = "", mega_state_code: str = "", queue_depth: Any = "", review_required: Any = "", travel_ms: Any = "", notes: str = "", raw_line: str = "", event_summary_tr: str = "", vision_event_id: Any = "", oee_snapshot_id: Any = "") -> dict[str, Any]:
        normalized_color = normalize_color(color_code)
        event_id_key = "vision_event" if event_type_code == "vision_event" else event_type_code
        return {
            "log_event_id": log_event_id,
            "event_time_text": received_at,
            "source_id": SOURCE_IDS.get(source_code, ""),
            "source_code": source_code,
            "event_type_id": EVENT_TYPE_IDS.get(event_id_key, ""),
            "event_type_code": event_type_code,
            "event_summary_tr": event_summary_tr,
            "item_id": item_id,
            "measure_id": measure_id,
            "fault_id": "",
            "oee_snapshot_id": oee_snapshot_id,
            "vision_event_id": vision_event_id,
            "line_id": 1,
            "station_id": _station_for_event(event_id_key),
            "color_id": _color_id(normalized_color),
            "color_code": normalized_color if normalized_color != "unknown" else "",
            "decision_source_id": _decision_source_id(decision_source_code),
            "decision_source_code": str(decision_source_code or "").strip().upper(),
            "mega_state_id": _mega_state_id(mega_state_code),
            "mega_state_code": mega_state_code,
            "queue_depth": _safe_int(queue_depth),
            "review_required": 1 if review_required is True else (0 if review_required is False else ""),
            "travel_ms": _safe_int(travel_ms),
            "notes": notes,
            "raw_line": raw_line,
        }

    def consume_local_counts_reset(self, received_at: str) -> dict[str, list[dict[str, Any]]]:
        return {
            "1_Olay_Logu": [self._event_row(log_event_id=self._next("log_event_id"), received_at=received_at, source_code="system", event_type_code="counts_reset", notes="UI sayac sifirlama islemi", raw_line="SYSTEM|COUNTS|RESET", event_summary_tr="UI uzerinden renk sayaclari sifirlandi")],
            RAW_LOG_SHEET_NAME: [self._raw_row(received_at=received_at, source_topic="local/system", source_code="system", raw_payload="SYSTEM|COUNTS|RESET", parsed_flag=1, event_type_code="counts_reset", notes="local_only")],
        }

    def consume_tablet_log(self, raw_line: str, received_at: str) -> dict[str, list[dict[str, Any]]]:
        rows = {RAW_LOG_SHEET_NAME: [self._raw_row(received_at=received_at, source_topic="sau/iot/mega/konveyor/tablet/log", source_code="tablet", raw_payload=raw_line, parsed_flag=0)]}
        parsed_oee = parse_tablet_oee_line(raw_line)
        if parsed_oee is not None:
            rows[RAW_LOG_SHEET_NAME][0].update({"parsed_flag": 1, "event_type_code": "tablet_oee_snapshot"})
            notes = ";".join(
                [
                    f"oee={parsed_oee['oee']}" if parsed_oee.get("oee") is not None else "",
                    f"kull={parsed_oee['availability']}" if parsed_oee.get("availability") is not None else "",
                    f"perf={parsed_oee['performance']}" if parsed_oee.get("performance") is not None else "",
                    f"kalite={parsed_oee['quality']}" if parsed_oee.get("quality") is not None else "",
                    f"toplam={parsed_oee['production']['total']}",
                ]
            ).strip(";")
            snapshot_id = self._next("oee_snapshot_id")
            event_log_id = self._next("log_event_id")
            rows["1_Olay_Logu"] = [
                self._event_row(
                    log_event_id=event_log_id,
                    received_at=received_at,
                    source_code="tablet",
                    event_type_code="tablet_oee_snapshot",
                    notes=notes,
                    raw_line=parsed_oee["raw_line"],
                    event_summary_tr="Tablet OEE ozeti alindi",
                    oee_snapshot_id=snapshot_id,
                )
            ]
            rows["5_OEE_Anliklari"] = [
                self._oee_snapshot_row(
                    snapshot_id=snapshot_id,
                    received_at=received_at,
                    event_log_id=event_log_id,
                    parsed_oee=parsed_oee,
                    raw_line=parsed_oee["raw_line"],
                    notes="source=tablet_snapshot",
                )
            ]
            return rows
        parsed_fault = parse_tablet_fault_line(raw_line)
        if parsed_fault is not None:
            rows[RAW_LOG_SHEET_NAME][0].update({"parsed_flag": 1, "event_type_code": "tablet_fault", "notes": parsed_fault["status"]})
            note_parts = [f"durum={parsed_fault['status']}", f"neden={parsed_fault['reason']}"]
            if parsed_fault.get("duration_ms") is not None:
                note_parts.append(f"sure_ms={parsed_fault['duration_ms']}")
            if parsed_fault.get("duration_min") is not None:
                note_parts.append(f"sure_dk={parsed_fault['duration_min']}")
            rows["1_Olay_Logu"] = [
                self._event_row(
                    log_event_id=self._next("log_event_id"),
                    received_at=received_at,
                    source_code="tablet",
                    event_type_code="tablet_fault",
                    notes=";".join(note_parts),
                    raw_line=parsed_fault["raw_line"],
                    event_summary_tr=f"Tablet ariza durumu: {_fault_status_label_tr(parsed_fault['status'])}",
                )
            ]
        return rows

    def consume_system_oee_log(self, raw_line: str, received_at: str) -> dict[str, list[dict[str, Any]]]:
        rows = {
            RAW_LOG_SHEET_NAME: [
                self._raw_row(
                    received_at=received_at,
                    source_topic="local/oee",
                    source_code="system",
                    raw_payload=raw_line,
                    parsed_flag=0,
                )
            ]
        }
        text = str(raw_line or "").strip()
        event_type_code = ""
        notes = ""
        summary = ""
        if text.startswith("SYSTEM|OEE|"):
            parts = [part.strip() for part in text.split("|")]
            action = parts[2] if len(parts) > 2 else ""
            value = parts[3] if len(parts) > 3 else ""
            action_code = normalize_token(action)
            event_type_code = "oee_control"
            notes = f"action={action_code}"
            if value:
                notes = f"{notes};value={value}"
            summary_map = {
                "select_shift": f"Vardiya secimi guncellendi: {value}",
                "set_performance_mode": f"Performans modu guncellendi: {value}",
                "set_target_qty": f"Hedef guncellendi: {value}",
                "set_ideal_cycle_sec": f"Ideal cevrim guncellendi: {value}",
                "set_planned_stop_min": f"Planli durus guncellendi: {value}",
            }
            summary = summary_map.get(action_code, f"OEE kontrol aksiyonu: {action_code}")
        elif "|Tablet|Sistem|" in text:
            _, fields = parse_colon_value_line(text, min_parts=2)
            event_name = normalize_token(str(fields.get("OLAY") or ""))
            shift_code = str(fields.get("VARDIYA") or "").strip()
            if event_name == "vardiya_basladi":
                event_type_code = "shift_start"
                notes = ";".join(
                    [
                        f"shift={shift_code}" if shift_code else "",
                        f"perf_mod={fields.get('PERF_MOD')}" if fields.get("PERF_MOD") else "",
                        f"hedef={fields.get('HEDEF')}" if fields.get("HEDEF") else "",
                        (
                            f"ideal_cycle_ms={int(round(float(fields.get('IDEAL_CYCLE_SN')) * 1000.0))}"
                            if fields.get("IDEAL_CYCLE_SN")
                            else ""
                        ),
                        f"ideal_cycle_sn={fields.get('IDEAL_CYCLE_SN')}" if fields.get("IDEAL_CYCLE_SN") else "",
                        (
                            f"planned_stop_ms={int(round(float(fields.get('PLANLI_DURUS_DK')) * 60_000.0))}"
                            if fields.get("PLANLI_DURUS_DK")
                            else ""
                        ),
                        f"planned_stop_dk={fields.get('PLANLI_DURUS_DK')}" if fields.get("PLANLI_DURUS_DK") else "",
                    ]
                ).strip(";")
                summary = f"Vardiya basladi: {shift_code}" if shift_code else "Vardiya basladi"
            elif event_name == "vardiya_bitti":
                event_type_code = "shift_stop"
                notes = ";".join(
                    [
                        f"shift={shift_code}" if shift_code else "",
                        f"toplam={fields.get('TOPLAM')}" if fields.get("TOPLAM") else "",
                        f"saglam={fields.get('SAGLAM')}" if fields.get("SAGLAM") else "",
                        f"rework={fields.get('REWORK')}" if fields.get("REWORK") else "",
                        f"hurda={fields.get('HURDA')}" if fields.get("HURDA") else "",
                    ]
                ).strip(";")
                summary = f"Vardiya bitti: {shift_code}" if shift_code else "Vardiya bitti"
        if not event_type_code:
            return rows
        rows[RAW_LOG_SHEET_NAME][0].update({"parsed_flag": 1, "event_type_code": event_type_code, "notes": notes})
        rows["1_Olay_Logu"] = [
            self._event_row(
                log_event_id=self._next("log_event_id"),
                received_at=received_at,
                source_code="system",
                event_type_code=event_type_code,
                notes=notes,
                raw_line=text,
                event_summary_tr=summary or event_type_code,
            )
        ]
        return rows

    def consume_kiosk_event(self, payload: Any, received_at: str) -> dict[str, list[dict[str, Any]]]:
        row = payload if isinstance(payload, dict) else {}
        event_type = str(row.get("event_type") or row.get("eventType") or "").strip()
        if not event_type:
            return {}
        item_id = str(row.get("item_id") or row.get("itemId") or "").strip()
        measure_id = str(row.get("measure_id") or row.get("measureId") or "").strip()
        note_parts = [
            f"device_id={row.get('device_id')}" if row.get("device_id") else "",
            f"station_id={row.get('bound_station_id')}" if row.get("bound_station_id") else "",
            f"operator_id={row.get('operator_id')}" if row.get("operator_id") else "",
            f"session_id={row.get('session_id')}" if row.get("session_id") else "",
            f"step_code={row.get('step_code')}" if row.get("step_code") else "",
            f"phase={row.get('phase')}" if row.get("phase") else "",
            f"fault_code={row.get('fault_code')}" if row.get("fault_code") else "",
            f"repeat_count={row.get('repeat_count')}" if row.get("repeat_count") not in (None, "") else "",
            f"status={row.get('status')}" if row.get("status") else "",
            f"note={row.get('note')}" if row.get("note") else "",
            f"reason={row.get('reason')}" if row.get("reason") else "",
        ]
        summary_map = {
            "maintenance_opening_started": "Acilis bakimi baslatildi",
            "maintenance_step_completed": f"Bakim adimi tamamlandi: {row.get('step_label') or row.get('step_code') or '-'}",
            "maintenance_completed": f"Bakim tamamlandi: {row.get('phase') or '-'}",
            "maintenance_closing_started": "Kapanis bakimi baslatildi",
            "kiosk_fault_started": f"Kiosk arizasi baslatildi: {row.get('reason') or row.get('fault_code') or '-'}",
            "kiosk_fault_cleared": "Kiosk arizasi kapatildi",
            "help_requested": "Teknisyen yardim istegi acildi",
            "help_acknowledged": "Yardim istegi onaylandi",
            "help_resolved": "Yardim istegi cozuldu",
        }
        raw_payload = {
            "event_type": event_type,
            "device_id": row.get("device_id") or "",
            "bound_station_id": row.get("bound_station_id") or "",
            "operator_id": row.get("operator_id") or "",
            "session_id": row.get("session_id") or "",
            "step_code": row.get("step_code") or "",
            "phase": row.get("phase") or "",
            "fault_code": row.get("fault_code") or "",
            "reason": row.get("reason") or "",
        }
        return {
            "1_Olay_Logu": [
                self._event_row(
                    log_event_id=self._next("log_event_id"),
                    received_at=received_at,
                    source_code="tablet",
                    event_type_code=event_type,
                    item_id=item_id,
                    measure_id=measure_id,
                    notes=";".join(part for part in note_parts if part),
                    raw_line=_json_text(raw_payload),
                    event_summary_tr=summary_map.get(event_type, event_type),
                )
            ],
            RAW_LOG_SHEET_NAME: [
                self._raw_row(
                    received_at=received_at,
                    source_topic="local/kiosk",
                    source_code="tablet",
                    raw_payload=raw_payload,
                    parsed_flag=1,
                    event_type_code=event_type,
                    item_id=item_id,
                    measure_id=measure_id,
                )
            ],
        }

    def consume_mega_log(self, raw_line: str, received_at: str) -> dict[str, list[dict[str, Any]]]:
        self.pending_completed_row_update = None
        rows = {RAW_LOG_SHEET_NAME: [self._raw_row(received_at=received_at, source_topic="sau/iot/mega/konveyor/logs", source_code="mega", raw_payload=raw_line, parsed_flag=0)]}
        parsed = parse_mega_event_from_log(raw_line)
        if parsed is None:
            return rows
        parsed, key = self._resolve_robot_event_context(parsed)
        rows[RAW_LOG_SHEET_NAME][0].update({"parsed_flag": 1, "event_type_code": parsed["event_type"], "item_id": parsed["item_id"], "measure_id": parsed["measure_id"], "color_code": parsed["color"]})
        log_event_id = self._next("log_event_id")
        raw = parsed["raw"]
        summary = {
            "measurement_decision": f"Olcum karari verildi: {_color_label_tr(parsed['color'])}",
            "queue_enq": f"Urun kuyruga alindi: {_color_label_tr(parsed['color'])}",
            "arm_position_reached": "Robot kol hedefe ulasti",
            "pickplace_started": "Robot tasima cevrimini baslatti",
            "pick_drop_reached": "Robot birakma noktasina ulasti",
            "pick_command_rejected": "Erken pick komutu reddedildi",
            "pick_released": "Urun birakildi",
            "pick_return_started": "Robot geri donuse basladi",
            "pick_return_reached": "Robot referans noktasina ulasti",
            "pickplace_done": "Birakma islemi tamamlandi",
            "pickplace_return_done": "Robot hazir bekleme konumuna dondu",
        }.get(parsed["event_type"], parsed["event_type"])
        note_parts: list[str] = []
        if parsed["event_type"] == "measurement_decision" and raw.get("SEARCH_HINT"):
            note_parts.append(f"search_hint={raw.get('SEARCH_HINT')}")
        if parsed["event_type"] == "measurement_decision" and raw.get("CONF") not in (None, ""):
            note_parts.append(f"conf={raw.get('CONF')}")
        if parsed.get("trigger_source") not in (None, "", "unknown"):
            note_parts.append(f"trigger={parsed['trigger_source']}")
        if parsed.get("reject_reason") not in (None, "", "unknown"):
            note_parts.append(f"reason={parsed['reject_reason']}")
        rows["1_Olay_Logu"] = [
            self._event_row(
                log_event_id=log_event_id,
                received_at=received_at,
                source_code="mega",
                event_type_code=parsed["event_type"],
                item_id=parsed["item_id"],
                measure_id=parsed["measure_id"],
                color_code=self._current_event_color(key, parsed),
                decision_source_code=self._current_event_decision_source(key, parsed),
                mega_state_code=parsed["mega_state"],
                queue_depth=parsed["queue_depth"],
                review_required=self._current_event_review_required(key, parsed),
                travel_ms=parsed["travel_ms"],
                notes=";".join(note_parts),
                raw_line=raw_line,
                event_summary_tr=summary,
            )
        ]

        if parsed["event_type"] == "measurement_decision":
            error_flag, error_reason = _measurement_error_info(parsed["color"], raw.get("CONF"))
            rows["2_Olcumler"] = [{
                "measurement_row_id": self._next("measurement_row_id"),
                "measure_id": parsed["measure_id"],
                "item_id": parsed["item_id"],
                "measured_at": received_at,
                "measurement_log_event_id": log_event_id,
                "source_log_file": "mqtt/logs",
                "final_color_id": _color_id(parsed["color"]),
                "final_color_code": parsed["color"],
                "final_color_raw": raw.get("FINAL", ""),
                "decision_source_id": _decision_source_id(parsed["decision_source"]),
                "decision_source_code": str(parsed["decision_source"] or "").upper(),
                "search_hint": str(raw.get("SEARCH_HINT") or ""),
                "search_hint_win": _safe_int(raw.get("SEARCH_HINT_WIN")),
                "search_hint_second": _safe_int(raw.get("SEARCH_HINT_SECOND")),
                "search_hint_strong": _safe_int(raw.get("SEARCH_HINT_STRONG")),
                "search_hint_fallback_allowed": _safe_int(raw.get("SEARCH_HINT_FALLBACK_ALLOWED")),
                "review_required": 1 if parsed["review_required"] else 0,
                "core_used": _safe_int(raw.get("CORE_USED")),
                "core_n": _safe_int(raw.get("CORE_N")),
                "obj_n": _safe_int(raw.get("OBJ_N")),
                "median_nearest": str(raw.get("MEDIAN_NEAREST") or ""),
                "score_nearest": str(raw.get("SCORE_NEAREST") or ""),
                "med_r": _safe_int(raw.get("MED_R")),
                "med_g": _safe_int(raw.get("MED_G")),
                "med_b": _safe_int(raw.get("MED_B")),
                "med_d_r": _safe_int(raw.get("MED_D_R")),
                "med_d_y": _safe_int(raw.get("MED_D_Y")),
                "med_d_b": _safe_int(raw.get("MED_D_B")),
                "med_d_x": _safe_int(raw.get("MED_D_X")),
                "x_r": _safe_int(raw.get("X_R")),
                "x_g": _safe_int(raw.get("X_G")),
                "x_b": _safe_int(raw.get("X_B")),
                "med_obj": _safe_int(raw.get("MED_OBJ")),
                "confidence": raw.get("CONF", ""),
                "core_str_min": _safe_int(raw.get("CORE_STR_MIN")),
                "core_str_max": _safe_int(raw.get("CORE_STR_MAX")),
                "vote_win": _safe_int(raw.get("VOTE_WIN")),
                "vote_y": _safe_int(raw.get("VOTE_Y")),
                "vote_second": _safe_int(raw.get("VOTE_SECOND")),
                "vote_classified": _safe_int(raw.get("VOTE_CLASSIFIED")),
                "vote_x": _safe_int(raw.get("VOTE_BOS")),
                "vote_r": _safe_int(raw.get("VOTE_R")),
                "vote_b": _safe_int(raw.get("VOTE_B")),
                "vote_cal": _safe_int(raw.get("VOTE_CAL")),
                "tot_r": _safe_int(raw.get("TOT_R")),
                "tot_y": _safe_int(raw.get("TOT_Y")),
                "tot_b": _safe_int(raw.get("TOT_B")),
                "tot_x": _safe_int(raw.get("TOT_BOS")),
                "tot_cal": _safe_int(raw.get("TOT_CAL")),
                "measurement_error_flag": error_flag,
                "measurement_error_reason": error_reason,
                "raw_line": raw_line,
            }]
            self.completed_state.setdefault(key, {}).update(
                {
                    "item_id": parsed["item_id"],
                    "measure_id": parsed["measure_id"],
                    "measured_at": received_at,
                    "detected_at": received_at,
                    "sensor_color": parsed["color"],
                    "final_color": parsed["color"],
                    "decision_source": parsed["decision_source"],
                    "sensor_decision_source": parsed["decision_source"],
                    "finalization_reason": "SENSOR_NO_VISION",
                    "review_required": parsed["review_required"],
                    "correlation_status": "",
                    "mismatch_flag": 0,
                    "early_pick_triggered": 0,
                    "pick_trigger_source": "",
                    "early_pick_request_sent_at": "",
                    "early_pick_accepted_at": "",
                }
            )
        elif parsed["event_type"] == "queue_enq":
            self.completed_state.setdefault(key, {}).update(
                {
                    "item_id": parsed["item_id"],
                    "measure_id": parsed["measure_id"],
                    "detected_at": self.completed_state.get(key, {}).get("detected_at", received_at),
                    "queued_at": received_at,
                    "sensor_color": parsed["color"],
                    "final_color": parsed["color"],
                    "travel_ms": parsed["travel_ms"],
                    "travel_ms_initial": parsed["travel_ms"],
                    "decision_source": parsed["decision_source"],
                    "sensor_decision_source": parsed["decision_source"],
                    "review_required": parsed["review_required"],
                    "queue_event_log_id": log_event_id,
                    "finalization_reason": "SENSOR_NO_VISION",
                    "correlation_status": "",
                    "mismatch_flag": 0,
                    "early_pick_triggered": 0,
                    "pick_trigger_source": "",
                    "early_pick_request_sent_at": "",
                    "early_pick_accepted_at": "",
                }
            )
        elif parsed["event_type"] == "arm_position_reached":
            state = self.completed_state.setdefault(key, {})
            trigger_source = str(parsed.get("trigger_source") or "").strip().upper()
            state.update(
                {
                    "pick_trigger_source": trigger_source or state.get("pick_trigger_source", ""),
                    "picked_at": received_at,
                }
            )
            if trigger_source == "EARLY":
                state["early_pick_accepted_at"] = received_at
                state["early_pick_triggered"] = 1
        elif parsed["event_type"] == "pick_command_rejected":
            state = self.completed_state.setdefault(key, {})
            state["reject_reason"] = str(parsed.get("reject_reason") or "").strip().upper()
        elif parsed["event_type"] == "pick_released":
            state = self.completed_state.setdefault(key, {})
            state["released_at"] = received_at
        elif parsed["event_type"] == "pick_return_started":
            self.completed_state.setdefault(key, {}).update({"return_started_at": received_at})
        elif parsed["event_type"] == "pick_return_reached":
            self.completed_state.setdefault(key, {}).update({"return_reached_at": received_at})
        elif parsed["event_type"] == "pickplace_done":
            state = self.completed_state.get(key, {})
            completed_key = str(state.get("item_id", parsed["item_id"]) or key)
            if state and completed_key not in self.completed_rows_by_item:
                rows["4_Uretim_Tamamlanan"] = [
                    self._finalize_completed_row(state=state, parsed=parsed, key=key, received_at=received_at, log_event_id=log_event_id)
                ]
        return rows

    def apply_quality_override(self, item_id: str, classification: str, applied_at: str) -> dict[str, Any]:
        normalized_item_id = str(item_id or "").strip()
        target = self.completed_rows_by_item.get(normalized_item_id)
        if not isinstance(target, dict):
            raise KeyError("ITEM_NOT_FOUND")
        normalized = str(classification or "").strip().upper()
        if normalized not in {"GOOD", "REWORK", "SCRAP"}:
            raise ValueError("INVALID_CLASSIFICATION")
        status_map = {
            "GOOD": ("COMPLETED", "Tamamlandi"),
            "REWORK": ("COMPLETED_REWORK", "Rework"),
            "SCRAP": ("COMPLETED_SCRAP", "Hurda"),
        }
        quality_map = {"GOOD": "Saglam", "REWORK": "Rework", "SCRAP": "Hurda"}
        status_code, status_tr = status_map[normalized]
        target["status_code"] = status_code
        target["status_tr"] = status_tr
        target["final_quality_code"] = normalized
        target["final_quality_tr"] = quality_map[normalized]
        target["override_flag"] = 1
        target["override_source_code"] = "MANUAL"
        target["override_applied_at"] = applied_at
        return target

    def consume_early_pick_request(self, item_id: str, received_at: str) -> dict[str, list[dict[str, Any]]]:
        normalized_item_id = str(item_id or "").strip()
        if not normalized_item_id:
            return {}
        self.pending_completed_row_update = None
        for state in self.completed_state.values():
            if str(state.get("item_id") or "").strip() != normalized_item_id:
                continue
            if not state.get("early_pick_request_sent_at"):
                state["early_pick_request_sent_at"] = received_at
            break
        return {
            "1_Olay_Logu": [
                self._event_row(
                    log_event_id=self._next("log_event_id"),
                    received_at=received_at,
                    source_code="system",
                    event_type_code="early_pick_request",
                    item_id=normalized_item_id,
                    notes="source=mes_web",
                    raw_line=f"SYSTEM|VISION|EARLY_PICK_REQUEST_SENT|ITEM_ID={normalized_item_id}",
                    event_summary_tr="Erken pick komutu gonderildi",
                )
            ],
            RAW_LOG_SHEET_NAME: [
                self._raw_row(
                    received_at=received_at,
                    source_topic="local/vision",
                    source_code="system",
                    raw_payload=f"SYSTEM|VISION|EARLY_PICK_REQUEST_SENT|ITEM_ID={normalized_item_id}",
                    parsed_flag=1,
                    event_type_code="early_pick_request",
                    item_id=normalized_item_id,
                )
            ],
        }

    def consume_vision_event(self, payload: Any, received_at: str) -> dict[str, list[dict[str, Any]]]:
        self.pending_completed_row_update = None
        rows = {RAW_LOG_SHEET_NAME: [self._raw_row(received_at=received_at, source_topic="sau/iot/mega/konveyor/vision/events", source_code="vision", raw_payload=payload, parsed_flag=0)]}
        parsed = parse_vision_event(payload)
        if parsed is None:
            return rows
        vision_event_id = self._next("vision_event_id")
        log_event_id = self._next("log_event_id")
        correlation_status = parsed.get("correlation_status") or ""
        translated_status = _correlation_status_label_tr(correlation_status) if correlation_status else ""
        raw_notes = [f"vision_event={_vision_event_label_tr(parsed['event_type'])}"]
        if translated_status:
            raw_notes.append(f"durum={translated_status}")
        if parsed["notes"]:
            raw_notes.append(parsed["notes"])
        rows[RAW_LOG_SHEET_NAME][0].update({"parsed_flag": 1, "event_type_code": "vision_event", "item_id": parsed.get("item_id") or "", "measure_id": parsed.get("measure_id") or "", "color_code": parsed["color"], "notes": parsed["event_type"]})
        rows["1_Olay_Logu"] = [
            self._event_row(
                log_event_id=log_event_id,
                received_at=received_at,
                source_code="vision",
                event_type_code="vision_event",
                item_id=parsed.get("item_id") or "",
                measure_id=parsed.get("measure_id") or "",
                color_code=parsed["color"],
                decision_source_code="VISION" if parsed.get("decision_applied") else "",
                review_required=parsed.get("review_required"),
                notes=";".join(part for part in raw_notes if part),
                raw_line=_json_text(payload),
                event_summary_tr=f"Vision olayi: {_vision_event_label_tr(parsed['event_type'])}",
                vision_event_id=vision_event_id,
            )
        ]
        raw = parsed["raw"]
        bbox = raw.get("bbox") or {}
        bbox_x1 = _safe_int(bbox.get("x1") if bbox.get("x1") is not None else bbox.get("x"))
        bbox_y1 = _safe_int(bbox.get("y1") if bbox.get("y1") is not None else bbox.get("y"))
        if bbox.get("x2") is not None:
            bbox_x2 = _safe_int(bbox.get("x2"))
        else:
            bbox_x2 = _safe_int((_safe_int(bbox.get("x")) or 0) + (_safe_int(bbox.get("w")) or 0))
        if bbox.get("y2") is not None:
            bbox_y2 = _safe_int(bbox.get("y2"))
        else:
            bbox_y2 = _safe_int((_safe_int(bbox.get("y")) or 0) + (_safe_int(bbox.get("h")) or 0))
        rows["6_Vision"] = [{
            "vision_event_id": vision_event_id,
            "event_time": received_at,
            "source_code": "vision",
            "vision_track_id": parsed["track_id"] or "",
            "event_type": parsed["event_type"],
            "color_id": _color_id(parsed["color"]),
            "color_code": parsed["color"],
            "item_id": parsed.get("item_id") or "",
            "measure_id": parsed.get("measure_id") or "",
            "confidence": parsed.get("confidence", ""),
            "confidence_tier": parsed.get("confidence_tier") or "",
            "correlation_status": parsed.get("correlation_status") or "",
            "late_vision_audit_flag": 1 if parsed.get("late_vision_audit_flag") else 0,
            "vision_observed_at": parsed.get("vision_observed_at") or "",
            "vision_published_at": parsed.get("vision_published_at") or "",
            "vision_received_at": received_at,
            "line_id": 1,
            "station_id": 6,
            "bbox_x1": bbox_x1,
            "bbox_y1": bbox_y1,
            "bbox_x2": bbox_x2,
            "bbox_y2": bbox_y2,
            "direction": raw.get("direction", ""),
            "is_placeholder": 0,
            "notes": parsed["notes"],
        }]
        key = self._completed_key(parsed.get("item_id"), parsed.get("measure_id"))
        if key in self.completed_state:
            state = self.completed_state[key]
            state.update(
                {
                    "vision_color": parsed["color"],
                    "vision_confidence": parsed.get("confidence", ""),
                    "vision_track_id": parsed.get("track_id") or "",
                    "vision_observed_at": parsed.get("vision_observed_at") or "",
                    "vision_published_at": parsed.get("vision_published_at") or "",
                    "vision_received_at": received_at,
                    "correlation_status": parsed.get("correlation_status") or state.get("correlation_status", ""),
                    "late_vision_audit_flag": 1 if parsed.get("late_vision_audit_flag") else 0,
                    "review_required": bool(parsed.get("review_required")) or bool(state.get("review_required")),
                }
            )
            if parsed.get("decision_applied"):
                state["final_color"] = parsed["color"]
                state["decision_source"] = "VISION"
                state["mismatch_flag"] = 1 if str(state.get("sensor_color") or "") != parsed["color"] else 0
                state["finalization_reason"] = "VISION_CORRECTED_MISMATCH" if state.get("mismatch_flag") else "VISION_HIGH_CONF"
        else:
            completed_row = self._find_completed_row(parsed.get("item_id"), parsed.get("measure_id"))
            if isinstance(completed_row, dict):
                self._set_completed_row_vision_fields(completed_row, parsed, received_at)
                self.pending_completed_row_update = completed_row
        return rows


class ExcelRuntimeSink:
    def __init__(self, config: AppConfig) -> None:
        self.config = config
        self.projector = WorkbookProjector()
        self._queue: queue.Queue[SinkEnvelope | None] = queue.Queue()
        self._thread: threading.Thread | None = None
        self._enabled = False

    def start(self) -> None:
        if self._enabled:
            return
        if not self.config.excel_enabled:
            return
        try:
            import openpyxl  # noqa: F401
        except ModuleNotFoundError:
            print("MES Web: Excel sink disabled because openpyxl is not installed.")
            return
        self._enabled = True
        self._thread = threading.Thread(target=self._worker, name="mes-web-excel-sink", daemon=True)
        self._thread.start()

    def stop(self) -> None:
        if not self._enabled:
            return
        self._queue.put(None)
        if self._thread is not None:
            self._thread.join(timeout=5)
        self._thread = None
        self._enabled = False

    def record_mega_log(self, raw_line: str, received_at: str) -> None:
        if self._enabled:
            self._queue.put(SinkEnvelope(kind="mega_log", received_at=received_at, payload=raw_line))

    def record_vision_event(self, payload: Any, received_at: str) -> None:
        if self._enabled:
            self._queue.put(SinkEnvelope(kind="vision_event", received_at=received_at, payload=payload))

    def record_tablet_log(self, raw_line: str, received_at: str) -> None:
        if self._enabled:
            self._queue.put(SinkEnvelope(kind="tablet_log", received_at=received_at, payload=raw_line))

    def record_system_oee_log(self, raw_line: str, received_at: str) -> None:
        if self._enabled:
            self._queue.put(SinkEnvelope(kind="system_oee_log", received_at=received_at, payload=raw_line))

    def record_local_counts_reset(self, received_at: str) -> None:
        if self._enabled:
            self._queue.put(SinkEnvelope(kind="counts_reset", received_at=received_at, payload="SYSTEM|COUNTS|RESET"))

    def record_quality_override(self, item_id: str, classification: str, received_at: str) -> None:
        if self._enabled:
            self._queue.put(
                SinkEnvelope(
                    kind="quality_override",
                    received_at=received_at,
                    payload={"item_id": item_id, "classification": classification},
                )
            )

    def record_early_pick_request(self, item_id: str, received_at: str) -> None:
        if self._enabled:
            self._queue.put(
                SinkEnvelope(
                    kind="early_pick_request",
                    received_at=received_at,
                    payload={"item_id": item_id},
                )
            )

    def record_work_order_state(self, state: dict[str, Any], received_at: str) -> None:
        if self._enabled:
            self._queue.put(
                SinkEnvelope(
                    kind="work_order_state",
                    received_at=received_at,
                    payload=state,
                )
            )

    def record_kiosk_event(self, event_type: str, payload: dict[str, Any], received_at: str) -> None:
        if self._enabled:
            envelope_payload = dict(payload or {})
            envelope_payload["event_type"] = str(event_type or "").strip()
            self._queue.put(
                SinkEnvelope(
                    kind="kiosk_event",
                    received_at=received_at,
                    payload=envelope_payload,
                )
            )

    def _worker(self) -> None:
        from openpyxl import Workbook, load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
        from openpyxl.utils import get_column_letter

        workbook_path = self.config.excel_workbook_path
        workbook = self._open_or_create_workbook(
            workbook_path=workbook_path,
            workbook_factory=Workbook,
            workbook_loader=load_workbook,
            invalid_file_error=InvalidFileException,
        )
        self._migrate_workbook_layout(workbook)
        if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) == 1 and workbook["Sheet"].max_row == 1 and workbook["Sheet"].max_column == 1 and workbook["Sheet"]["A1"].value is None:
            workbook.remove(workbook["Sheet"])
        counters: dict[str, int] = {}
        for sheet_name, headers in SHEET_COLUMNS.items():
            sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
            if sheet.max_row == 1 and all(sheet.cell(1, col).value is None for col in range(1, max(2, len(headers) + 1))):
                for col_index, header in enumerate(headers, start=1):
                    sheet.cell(1, col_index, header)
            elif sheet.cell(1, 1).value is None:
                for col_index, header in enumerate(headers, start=1):
                    sheet.cell(1, col_index, header)
            self._ensure_sheet_layout(sheet, headers, workbook=workbook)
            if sheet_name == RAW_LOG_SHEET_NAME:
                sheet.sheet_state = "hidden"
            id_header = headers[0]
            id_index = headers.index(id_header) + 1
            next_id = 1
            for row_index in range(2, sheet.max_row + 1):
                value = sheet.cell(row_index, id_index).value
                if isinstance(value, int):
                    next_id = max(next_id, value + 1)
            counters[id_header] = next_id
        self.projector.prime(counters)

        dirty = False
        while True:
            envelope = self._queue.get()
            if envelope is None:
                break
            batch = [envelope]
            while len(batch) < self.config.excel_batch_size:
                try:
                    next_envelope = self._queue.get(timeout=self.config.excel_flush_interval_sec)
                except queue.Empty:
                    break
                if next_envelope is None:
                    self._queue.put(None)
                    break
                batch.append(next_envelope)
            for item in batch:
                if item.kind == "mega_log":
                    rows = self.projector.consume_mega_log(item.payload, item.received_at)
                elif item.kind == "vision_event":
                    rows = self.projector.consume_vision_event(item.payload, item.received_at)
                    if self.projector.pending_completed_row_update is not None:
                        self._update_completed_sheet_row(workbook["4_Uretim_Tamamlanan"], self.projector.pending_completed_row_update)
                        self.projector.pending_completed_row_update = None
                        dirty = True
                elif item.kind == "tablet_log":
                    rows = self.projector.consume_tablet_log(item.payload, item.received_at)
                elif item.kind == "system_oee_log":
                    rows = self.projector.consume_system_oee_log(item.payload, item.received_at)
                elif item.kind == "quality_override":
                    row = self.projector.apply_quality_override(item.payload["item_id"], item.payload["classification"], item.received_at)
                    self._update_completed_sheet_row(workbook["4_Uretim_Tamamlanan"], row)
                    dirty = True
                    rows = {}
                elif item.kind == "early_pick_request":
                    rows = self.projector.consume_early_pick_request(item.payload["item_id"], item.received_at)
                elif item.kind == "work_order_state":
                    self._sync_work_order_sheets(workbook, item.payload, item.received_at)
                    self._sync_oee_snapshot_sheet(workbook, item.payload)
                    self._sync_fault_sheet(workbook, item.payload)
                    self._sync_maintenance_sheet(workbook, item.payload)
                    dirty = True
                    rows = {}
                elif item.kind == "kiosk_event":
                    rows = self.projector.consume_kiosk_event(item.payload, item.received_at)
                else:
                    rows = self.projector.consume_local_counts_reset(item.received_at)
                for sheet_name, row_dicts in rows.items():
                    sheet = workbook[sheet_name]
                    headers = [sheet.cell(1, idx).value for idx in range(1, sheet.max_column + 1)]
                    for row in row_dicts:
                        target_row = self._next_write_row(sheet, len(headers))
                        if target_row > 2:
                            self._copy_row_style(sheet, source_row=2, target_row=target_row, width=len(headers))
                        self._write_sheet_row(sheet, target_row, headers, row)
                        self._update_auto_filter(sheet, len(headers), target_row, get_column_letter)
                        dirty = True
            if dirty:
                workbook.save(workbook_path)
                dirty = False
        workbook.save(workbook_path)

    def _migrate_workbook_layout(self, workbook: Any) -> None:
        if LEGACY_WORK_ORDER_SHEET_NAME in workbook.sheetnames and WORK_ORDER_SHEET_NAME not in workbook.sheetnames:
            workbook[LEGACY_WORK_ORDER_SHEET_NAME].title = WORK_ORDER_SHEET_NAME
        if LEGACY_INVENTORY_SHEET_NAME in workbook.sheetnames and INVENTORY_SHEET_NAME not in workbook.sheetnames:
            workbook[LEGACY_INVENTORY_SHEET_NAME].title = INVENTORY_SHEET_NAME
        if LEGACY_RAW_LOG_SHEET_NAME in workbook.sheetnames and RAW_LOG_SHEET_NAME not in workbook.sheetnames:
            workbook[LEGACY_RAW_LOG_SHEET_NAME].title = RAW_LOG_SHEET_NAME
        if RAW_LOG_SHEET_NAME in workbook.sheetnames:
            raw_sheet = workbook[RAW_LOG_SHEET_NAME]
            raw_sheet.sheet_state = "hidden"
            if workbook.sheetnames[-1] != RAW_LOG_SHEET_NAME and hasattr(workbook, "_sheets"):
                workbook._sheets = [sheet for sheet in workbook._sheets if sheet.title != RAW_LOG_SHEET_NAME] + [raw_sheet]
        preferred_order = [
            "0_Tanimlamalar",
            "1_Olay_Logu",
            "2_Olcumler",
            "3_Arizalar",
            "4_Uretim_Tamamlanan",
            "5_OEE_Anliklari",
            "6_Vision",
            WORK_ORDER_SHEET_NAME,
            INVENTORY_SHEET_NAME,
            MAINTENANCE_SHEET_NAME,
            RAW_LOG_SHEET_NAME,
        ]
        if hasattr(workbook, "_sheets"):
            ordered_sheets = []
            seen_titles: set[str] = set()
            for sheet_name in preferred_order:
                if sheet_name in workbook.sheetnames and sheet_name not in seen_titles:
                    ordered_sheets.append(workbook[sheet_name])
                    seen_titles.add(sheet_name)
            for sheet in workbook.worksheets:
                if sheet.title in seen_titles:
                    continue
                ordered_sheets.append(sheet)
                seen_titles.add(sheet.title)
            workbook._sheets = ordered_sheets

    def _open_or_create_workbook(
        self,
        *,
        workbook_path: Path,
        workbook_factory: Any,
        workbook_loader: Any,
        invalid_file_error: type[Exception],
    ) -> Any:
        workbook_path.parent.mkdir(parents=True, exist_ok=True)
        template_path = self.config.excel_template_path
        if not workbook_path.exists() and template_path and template_path.exists():
            shutil.copyfile(template_path, workbook_path)
        try:
            return workbook_loader(workbook_path) if workbook_path.exists() else workbook_factory()
        except (OSError, zipfile.BadZipFile, invalid_file_error):
            archived_path = workbook_path.with_suffix(f"{workbook_path.suffix}.corrupt-{datetime.now().strftime('%Y%m%d-%H%M%S')}")
            try:
                if workbook_path.exists():
                    workbook_path.replace(archived_path)
            except OSError:
                archived_path = None
            if template_path and template_path.exists():
                try:
                    shutil.copyfile(template_path, workbook_path)
                    return workbook_loader(workbook_path)
                except (OSError, zipfile.BadZipFile, invalid_file_error):
                    pass
            workbook = workbook_factory()
            if archived_path is not None:
                print(f"MES Web: invalid workbook archived to {archived_path}")
            else:
                print(f"MES Web: invalid workbook ignored at {workbook_path}")
            return workbook

    def _sync_work_order_sheets(self, workbook: Any, state: dict[str, Any], received_at: str) -> None:
        work_orders = state.get("workOrders") if isinstance(state.get("workOrders"), dict) else {}
        orders = work_orders.get("ordersById") if isinstance(work_orders.get("ordersById"), dict) else {}
        sequence = work_orders.get("orderSequence") if isinstance(work_orders.get("orderSequence"), list) else []
        source = work_orders.get("source") if isinstance(work_orders.get("source"), dict) else {}
        source_file = str(source.get("file") or "")

        work_order_sheet = workbook[WORK_ORDER_SHEET_NAME]
        inventory_sheet = workbook[INVENTORY_SHEET_NAME]

        for order_id in sequence:
            order = orders.get(order_id)
            if not isinstance(order, dict):
                continue
            snapshot = build_work_order_snapshot(state, order)
            requirements = order.get("requirements") if isinstance(order.get("requirements"), list) else []
            requirement_note = ",".join(
                f"{str(req.get('color') or req.get('stockCode') or req.get('lineId') or '').strip()}:{int(req.get('completedQty') or 0)}/{int(req.get('quantity') or 0)}"
                for req in requirements
                if isinstance(req, dict)
            )
            target_id = str(order_id)
            row = {
                "work_order_record_id": self._existing_or_next_id(work_order_sheet, "work_order_record_id", "order_id", target_id),
                "order_id": target_id,
                "erp_type": str(order.get("erpType") or ""),
                "source_file": source_file,
                "queued_at": order.get("queuedAt"),
                "started_at": order.get("startedAt"),
                "completed_at": order.get("completedAt"),
                "status_code": str(order.get("status") or "").upper(),
                "status_tr": _work_order_status_tr(order.get("status")),
                "stock_code": str(order.get("stockCode") or ""),
                "stock_name": str(order.get("stockName") or ""),
                "product_color": str(order.get("productColor") or ""),
                "target_qty": int(snapshot["targetQty"]),
                "fulfilled_qty": int(snapshot["fulfilledQty"]),
                "production_qty": int(snapshot["productionQty"]),
                "inventory_consumed_qty": int(snapshot["inventoryConsumedQty"]),
                "good_qty": int(snapshot["goodQty"]),
                "rework_qty": int(snapshot["reworkQty"]),
                "scrap_qty": int(snapshot["scrapQty"]),
                "ideal_cycle_ms": int(snapshot["idealCycleMs"]),
                "ideal_cycle_sec": round(float(snapshot["idealCycleSec"]), 1),
                "planned_duration_ms": int(snapshot["plannedDurationMs"]),
                "planned_duration_sec": round(float(snapshot["plannedDurationMs"]) / 1000.0, 1),
                "runtime_ms": int(snapshot["runtimeMs"]),
                "runtime_sec": round(float(snapshot["runtimeMs"]) / 1000.0, 1),
                "unplanned_downtime_ms": int(snapshot["unplannedMs"]),
                "unplanned_downtime_sec": round(float(snapshot["unplannedMs"]) / 1000.0, 1),
                "availability_pct": _optional_pct(snapshot.get("availability")),
                "performance_pct": _optional_pct(snapshot.get("performance")),
                "quality_pct": _optional_pct(snapshot.get("quality")),
                "oee_pct": _optional_pct(snapshot.get("oee")),
                "started_by": str(order.get("startedBy") or ""),
                "started_by_name": str(order.get("startedByName") or ""),
                "transition_reason": str(order.get("transitionReason") or ""),
                "notes": f"updated_at={received_at}" + (f";requirements={requirement_note}" if requirement_note else ""),
            }
            self._upsert_sheet_row(work_order_sheet, "work_order_record_id", "order_id", target_id, row)

        inventory_rows = work_orders.get("inventoryByProduct") if isinstance(work_orders.get("inventoryByProduct"), dict) else {}
        for match_key, entry in inventory_rows.items():
            if not isinstance(entry, dict):
                continue
            target_key = str(match_key or "")
            if not target_key:
                continue
            row = {
                "inventory_record_id": self._existing_or_next_id(inventory_sheet, "inventory_record_id", "match_key", target_key),
                "match_key": target_key,
                "product_code": str(entry.get("productCode") or ""),
                "stock_code": str(entry.get("stockCode") or ""),
                "stock_name": str(entry.get("stockName") or ""),
                "color_code": str(entry.get("color") or ""),
                "quantity": int(entry.get("quantity") or 0),
                "last_updated_at": entry.get("lastUpdatedAt"),
                "last_source": str(entry.get("lastSource") or ""),
                "source_file": source_file,
                "notes": f"updated_at={received_at}",
            }
            self._upsert_sheet_row(inventory_sheet, "inventory_record_id", "match_key", target_key, row)
        self._zero_missing_inventory_rows(inventory_sheet, set(inventory_rows.keys()), received_at)
        self._update_auto_filter(work_order_sheet, work_order_sheet.max_column, max(work_order_sheet.max_row, 2), None)
        self._update_auto_filter(inventory_sheet, inventory_sheet.max_column, max(inventory_sheet.max_row, 2), None)

    def _sync_oee_snapshot_sheet(self, workbook: Any, state: dict[str, Any]) -> None:
        trend_rows = state.get("trend") if isinstance(state.get("trend"), list) else []
        if not trend_rows:
            return
        snapshot_sheet = workbook["5_OEE_Anliklari"]
        for row in trend_rows:
            if not isinstance(row, dict):
                continue
            snapshot_time = str(row.get("time") or "").strip()
            if not snapshot_time:
                continue
            reason = str(row.get("reason") or "").strip().lower()
            summary = str(row.get("summary") or "").strip()
            note_parts = [f"reason={reason}" if reason else "reason=runtime_state"]
            if summary:
                note_parts.append(f"summary={summary}")
            target_row = {
                "oee_snapshot_id": self._existing_or_next_id(snapshot_sheet, "oee_snapshot_id", "snapshot_time", snapshot_time),
                "snapshot_time": snapshot_time,
                "event_log_id": "",
                "sample_cycle_tag": f"runtime:{reason or 'snapshot'}:{snapshot_time}",
                "oee": row.get("oee"),
                "availability": row.get("availability"),
                "performance": row.get("performance"),
                "quality": row.get("quality"),
                "mavi_s": _safe_int(row.get("mavi_s")),
                "mavi_r": _safe_int(row.get("mavi_r")),
                "mavi_h": _safe_int(row.get("mavi_h")),
                "sari_s": _safe_int(row.get("sari_s")),
                "sari_r": _safe_int(row.get("sari_r")),
                "sari_h": _safe_int(row.get("sari_h")),
                "kirmizi_s": _safe_int(row.get("kirmizi_s")),
                "kirmizi_r": _safe_int(row.get("kirmizi_r")),
                "kirmizi_h": _safe_int(row.get("kirmizi_h")),
                "is_full_cycle_reference": 0 if reason == "periodic_30s" else 1,
                "notes": ";".join(part for part in note_parts if part),
                "raw_line": "",
            }
            self._upsert_sheet_row(snapshot_sheet, "oee_snapshot_id", "snapshot_time", snapshot_time, target_row)
        self._update_auto_filter(snapshot_sheet, snapshot_sheet.max_column, max(snapshot_sheet.max_row, 2), None)

    def _sync_fault_sheet(self, workbook: Any, state: dict[str, Any]) -> None:
        sheet = workbook["3_Arizalar"]
        payload_rows: list[dict[str, Any]] = []
        active_fault = state.get("activeFault") if isinstance(state.get("activeFault"), dict) else None
        if isinstance(active_fault, dict) and str(active_fault.get("source") or "").strip().lower() == "kiosk":
            payload_rows.append(dict(active_fault))
        for row in state.get("faultHistory") if isinstance(state.get("faultHistory"), list) else []:
            if not isinstance(row, dict) or str(row.get("source") or "").strip().lower() != "kiosk":
                continue
            payload_rows.append(dict(row))
        for row in payload_rows:
            fault_id = str(row.get("faultId") or row.get("fault_id") or "").strip()
            if not fault_id:
                continue
            started_at = row.get("startedAt") or row.get("started_at") or ""
            ended_at = row.get("endedAt") or row.get("ended_at") or ""
            duration_ms = int(float(row.get("durationMs") or 0))
            duration_sec = round(duration_ms / 1000.0, 1)
            status_code = "AKTIF" if not ended_at else "BITTI"
            target_row = {
                "fault_record_id": self._existing_or_next_id(sheet, "fault_record_id", "fault_id", fault_id),
                "fault_id": fault_id,
                "status_code": status_code,
                "status_tr": "Aktif" if status_code == "AKTIF" else "Bitti",
                "fault_type_code": str(row.get("reasonCode") or row.get("faultTypeCode") or "").strip(),
                "fault_category": str(row.get("category") or "").strip(),
                "fault_reason_tr": str(row.get("reason") or "").strip(),
                "started_at": started_at,
                "ended_at": ended_at,
                "duration_ms": duration_ms,
                "duration_sec": duration_sec,
                "source_code": str(row.get("source") or "kiosk").strip(),
                "device_id": str(row.get("deviceId") or row.get("device_id") or "").strip(),
                "device_name": str(row.get("deviceName") or row.get("device_name") or "").strip(),
                "bound_station_id": str(row.get("boundStationId") or row.get("bound_station_id") or "").strip(),
                "operator_id": str(row.get("operatorId") or row.get("operator_id") or "").strip(),
                "operator_code": str(row.get("operatorCode") or row.get("operator_code") or "").strip(),
                "operator_name": str(row.get("operatorName") or row.get("operator_name") or "").strip(),
                "shift_code": str(((state.get("shift") or {}) if isinstance(state.get("shift"), dict) else {}).get("code") or "").strip(),
                "notes": "",
            }
            self._upsert_sheet_row(sheet, "fault_record_id", "fault_id", fault_id, target_row)
        self._update_auto_filter(sheet, sheet.max_column, max(sheet.max_row, 2), None)

    def _sync_maintenance_sheet(self, workbook: Any, state: dict[str, Any]) -> None:
        sheet = workbook[MAINTENANCE_SHEET_NAME]
        maintenance = state.get("maintenance") if isinstance(state.get("maintenance"), dict) else {}
        sessions: list[dict[str, Any]] = []
        for key in ("openingSession", "closingSession"):
            session = maintenance.get(key)
            if isinstance(session, dict):
                sessions.append(session)
        for session in maintenance.get("history") if isinstance(maintenance.get("history"), list) else []:
            if isinstance(session, dict):
                sessions.append(session)
        for session in sessions:
            session_id = str(session.get("sessionId") or session.get("session_id") or "").strip()
            if not session_id:
                continue
            phase_code = str(session.get("phase") or "").strip().lower() or "opening"
            phase_tr = "Acilis" if phase_code == "opening" else "Kapanis"
            session_started_at = session.get("startedAt") or session.get("started_at") or ""
            session_ended_at = session.get("endedAt") or session.get("ended_at") or ""
            duration_ms = _duration_milliseconds(session_started_at, session_ended_at)
            if duration_ms is None:
                duration_ms = int(float(session.get("durationMs") or 0))
            duration_sec = round(duration_ms / 1000.0, 1)
            for index, step in enumerate(session.get("steps") if isinstance(session.get("steps"), list) else [], start=1):
                if not isinstance(step, dict):
                    continue
                step_code = str(step.get("stepCode") or step.get("step_code") or f"{phase_code}_step_{index}").strip()
                row_key = f"{session_id}:{step_code}"
                target_row = {
                    "maintenance_record_id": self._existing_or_next_id(sheet, "maintenance_record_id", "maintenance_row_key", row_key),
                    "maintenance_row_key": row_key,
                    "session_id": session_id,
                    "phase_code": phase_code,
                    "phase_tr": phase_tr,
                    "step_code": step_code,
                    "step_label_tr": str(step.get("stepLabel") or step.get("step_label") or step_code).strip(),
                    "required_flag": 1 if bool(step.get("required", True)) else 0,
                    "completed_flag": 1 if bool(step.get("completed")) else 0,
                    "completed_at": step.get("completedAt") or step.get("completed_at") or "",
                    "session_started_at": session_started_at,
                    "session_ended_at": session_ended_at,
                    "duration_ms": duration_ms,
                    "duration_sec": duration_sec,
                    "shift_code": str(session.get("shiftCode") or session.get("shift_code") or "").strip(),
                    "device_id": str(session.get("deviceId") or session.get("device_id") or "").strip(),
                    "device_name": str(session.get("deviceName") or session.get("device_name") or "").strip(),
                    "device_role": str(session.get("deviceRole") or session.get("device_role") or "").strip(),
                    "bound_station_id": str(session.get("boundStationId") or session.get("bound_station_id") or "").strip(),
                    "operator_id": str(session.get("operatorId") or session.get("operator_id") or "").strip(),
                    "operator_code": str(session.get("operatorCode") or session.get("operator_code") or "").strip(),
                    "operator_name": str(session.get("operatorName") or session.get("operator_name") or "").strip(),
                    "note": str(session.get("note") or "").strip(),
                }
                self._upsert_sheet_row(sheet, "maintenance_record_id", "maintenance_row_key", row_key, target_row)
        self._update_auto_filter(sheet, sheet.max_column, max(sheet.max_row, 2), None)

    def _existing_or_next_id(self, sheet: Any, id_header: str, key_header: str, target_key: str) -> int:
        headers = [sheet.cell(1, idx).value for idx in range(1, sheet.max_column + 1)]
        if not headers:
            return 1
        key_index = headers.index(key_header) + 1
        id_index = headers.index(id_header) + 1
        max_id = 0
        for row_index in range(2, sheet.max_row + 1):
            row_key = sheet.cell(row_index, key_index).value
            row_id = sheet.cell(row_index, id_index).value
            if isinstance(row_id, int):
                max_id = max(max_id, row_id)
            if str(row_key or "").strip() == target_key:
                return int(row_id or max_id + 1 or 1)
        return max_id + 1 if max_id > 0 else 1

    def _upsert_sheet_row(self, sheet: Any, id_header: str, key_header: str, target_key: str, row: dict[str, Any]) -> None:
        headers = [sheet.cell(1, idx).value for idx in range(1, sheet.max_column + 1)]
        if not headers:
            return
        key_index = headers.index(key_header) + 1
        target_row = None
        for row_index in range(2, sheet.max_row + 1):
            if str(sheet.cell(row_index, key_index).value or "").strip() == target_key:
                target_row = row_index
                break
        if target_row is None:
            target_row = self._next_write_row(sheet, len(headers))
            if target_row > 2:
                self._copy_row_style(sheet, source_row=2, target_row=target_row, width=len(headers))
        self._write_sheet_row(sheet, target_row, headers, row)

    def _zero_missing_inventory_rows(self, sheet: Any, active_keys: set[str], received_at: str) -> None:
        headers = [sheet.cell(1, idx).value for idx in range(1, sheet.max_column + 1)]
        if not headers or "match_key" not in headers:
            return
        key_index = headers.index("match_key") + 1
        quantity_index = headers.index("quantity") + 1 if "quantity" in headers else None
        notes_index = headers.index("notes") + 1 if "notes" in headers else None
        for row_index in range(2, sheet.max_row + 1):
            row_key = str(sheet.cell(row_index, key_index).value or "").strip()
            if not row_key or row_key in active_keys:
                continue
            if quantity_index is not None:
                sheet.cell(row_index, quantity_index, _excel_cell_value(0))
            if notes_index is not None:
                sheet.cell(row_index, notes_index, _excel_cell_value(f"updated_at={received_at};depleted=1"))

    def _update_completed_sheet_row(self, sheet: Any, row: dict[str, Any]) -> None:
        headers = [sheet.cell(1, idx).value for idx in range(1, sheet.max_column + 1)]
        if not headers:
            return
        id_index = headers.index("production_record_id") + 1
        target_id = row.get("production_record_id")
        target_row = None
        for row_index in range(2, sheet.max_row + 1):
            if sheet.cell(row_index, id_index).value == target_id:
                target_row = row_index
                break
        if target_row is None:
            raise KeyError("WORKBOOK_COMPLETED_ROW_NOT_FOUND")
        self._write_sheet_row(sheet, target_row, headers, row)

    def _write_sheet_row(self, sheet: Any, target_row: int, headers: list[str], row: dict[str, Any]) -> None:
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(target_row, col_index).value = _excel_cell_value(row.get(header, ""))

    def _ensure_sheet_layout(self, sheet: Any, headers: list[str], workbook: Any | None = None) -> None:
        existing_headers = [sheet.cell(1, idx).value for idx in range(1, max(sheet.max_column, 1) + 1)]
        existing_lookup = {str(value): idx for idx, value in enumerate(existing_headers, start=1) if value not in (None, "")}
        next_column = max(sheet.max_column, 0) + 1
        for header in headers:
            if header in existing_lookup:
                continue
            sheet.cell(1, next_column, header)
            next_column += 1
        if workbook is not None:
            self._apply_reference_header_style(workbook, sheet, len(headers))
        if sheet.freeze_panes is None:
            sheet.freeze_panes = "A2"
        if sheet.max_row < 2:
            sheet.append(["" for _ in headers])
        self._update_auto_filter(sheet, sheet.max_column, max(sheet.max_row, 2), None)

    def _apply_reference_header_style(self, workbook: Any, sheet: Any, width: int) -> None:
        if width <= 0 or not self._header_row_needs_style(sheet, width):
            return
        reference_cell = self._reference_header_style_cell(workbook, skip_sheet=sheet.title)
        if reference_cell is None or not reference_cell.has_style:
            return
        for col_index in range(1, width + 1):
            target = sheet.cell(1, col_index)
            target._style = copy(reference_cell._style)
            target.font = copy(reference_cell.font)
            target.fill = copy(reference_cell.fill)
            target.border = copy(reference_cell.border)
            target.alignment = copy(reference_cell.alignment)
            target.protection = copy(reference_cell.protection)
            target.number_format = reference_cell.number_format

    def _reference_header_style_cell(self, workbook: Any, *, skip_sheet: str = "") -> Any | None:
        preferred_sheet_names = [
            "1_Olay_Logu",
            "2_Olcumler",
            "3_Arizalar",
            "4_Uretim_Tamamlanan",
            "5_OEE_Anliklari",
            "6_Vision",
            MAINTENANCE_SHEET_NAME,
            RAW_LOG_SHEET_NAME,
        ]
        for sheet_name in preferred_sheet_names + list(workbook.sheetnames):
            if sheet_name == skip_sheet or sheet_name not in workbook.sheetnames:
                continue
            reference_sheet = workbook[sheet_name]
            for col_index in range(1, reference_sheet.max_column + 1):
                cell = reference_sheet.cell(1, col_index)
                if cell.has_style and cell.value not in (None, ""):
                    return cell
        return None

    def _header_row_needs_style(self, sheet: Any, width: int) -> bool:
        for col_index in range(1, width + 1):
            if sheet.cell(1, col_index).has_style:
                return False
        return True

    def _next_write_row(self, sheet: Any, width: int) -> int:
        for row_index in range(2, max(sheet.max_row, 2) + 1):
            if all(sheet.cell(row_index, col_index).value in (None, "") for col_index in range(1, width + 1)):
                return row_index
        return sheet.max_row + 1

    def _copy_row_style(self, sheet: Any, *, source_row: int, target_row: int, width: int) -> None:
        for col_index in range(1, width + 1):
            source = sheet.cell(source_row, col_index)
            target = sheet.cell(target_row, col_index)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.font:
                target.font = copy(source.font)
            if source.fill:
                target.fill = copy(source.fill)
            if source.border:
                target.border = copy(source.border)
            if source.alignment:
                target.alignment = copy(source.alignment)
            if source.protection:
                target.protection = copy(source.protection)

    def _update_auto_filter(self, sheet: Any, width: int, row_index: int, get_column_letter: Any | None) -> None:
        if get_column_letter is None:
            from openpyxl.utils import get_column_letter as local_get_column_letter

            get_column_letter = local_get_column_letter
        sheet.auto_filter.ref = f"A1:{get_column_letter(width)}{max(row_index, 2)}"
