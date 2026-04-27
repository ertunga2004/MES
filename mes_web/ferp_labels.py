from __future__ import annotations

import contextlib
import re
from dataclasses import asdict, dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover - environment-specific dependency gap.
    load_workbook = None


FERP_LABELS_RELATIVE_PATH = Path("README") / "ferp_labels.xlsx"

SUPPORTED_OBJECTS = {
    "mym4104",
    "mym4004",
    "mym4008",
    "mym4009",
    "mym4043",
    "mym4086",
    "mym2008",
    "mym2010",
    "mym2056",
}

REQUIRED_LABELS_BY_OBJECT = {
    "mym4104": {"lblMMFB0_NUMBER", "lblMMFB0_QTY"},
    "mym4004": {"lblMMFB0_NUMBER", "lblMMFB0_QTY"},
    "mym4008": {"lblMMFB0_NUMBER", "lblMMFB0_QTY"},
    "mym4009": {"lblMMFB0_NUMBER", "lblMMFB0_QTY"},
    "mym4043": {"lblMMFB0_NUMBER", "lblMMFB0_QTY"},
    "mym4086": {"lblMMFB0_NUMBER", "lblMMFB0_QTY"},
    "mym2008": {"lblMMV00_DATE", "lblMMV00_REF_NUMBER"},
    "mym2010": {"lblMMV00_DATE", "lblMMV00_REF_NUMBER"},
    "mym2056": {"lblMMV00_DATE", "lblMMV00_REF_NUMBER", "lblMWR00_CODE_O", "lblMWR00_CODE_I"},
}


class FerpLabelRegistryError(RuntimeError):
    pass


@dataclass(frozen=True, slots=True)
class FerpLabel:
    ferp_module: str
    ferp_screen: str
    ferp_object: str
    label_code: str
    label_text: str

    def to_dict(self) -> dict[str, str]:
        return asdict(self)


def default_ferp_labels_path(root_dir: str | Path | None = None) -> Path:
    base = Path(root_dir) if root_dir is not None else Path(__file__).resolve().parent.parent
    return base / FERP_LABELS_RELATIVE_PATH


def _cell_text(value: Any) -> str:
    return str(value or "").strip()


def _is_object_code(value: Any) -> bool:
    return bool(re.fullmatch(r"mym\d+", _cell_text(value).lower()))


def _is_label_code(value: Any) -> bool:
    return _cell_text(value).startswith("lbl")


def _nearest_text_left(ws: Any, row: int, col: int) -> str:
    for candidate_col in range(col - 1, 0, -1):
        value = _cell_text(ws.cell(row, candidate_col).value)
        if value and not _is_label_code(value) and not _is_object_code(value):
            return value
    return ""


def _label_text_for_cell(ws: Any, row: int, col: int) -> str:
    return _nearest_text_left(ws, row, col)


def _extract_sheet_registry(ws: Any) -> dict[str, dict[str, FerpLabel]]:
    object_rows: list[tuple[int, int, str, str]] = []
    for row in ws.iter_rows():
        for cell in row:
            value = _cell_text(cell.value)
            if _is_object_code(value):
                object_rows.append((int(cell.row), int(cell.column), value.lower(), _nearest_text_left(ws, cell.row, cell.column)))

    object_rows.sort(key=lambda item: (item[0], item[1]))
    registry: dict[str, dict[str, FerpLabel]] = {}
    for index, (row_index, _col_index, object_code, screen_name) in enumerate(object_rows):
        next_row = object_rows[index + 1][0] if index + 1 < len(object_rows) else ws.max_row + 1
        if object_code not in SUPPORTED_OBJECTS:
            continue
        labels: dict[str, FerpLabel] = registry.setdefault(object_code, {})
        for scan_row in range(row_index + 1, next_row):
            for scan_col in range(1, ws.max_column + 1):
                label_code = _cell_text(ws.cell(scan_row, scan_col).value)
                if not _is_label_code(label_code) or label_code in labels:
                    continue
                labels[label_code] = FerpLabel(
                    ferp_module=str(ws.title),
                    ferp_screen=screen_name,
                    ferp_object=object_code,
                    label_code=label_code,
                    label_text=_label_text_for_cell(ws, scan_row, scan_col),
                )
    return registry


def _validate_source_path(path: Path) -> tuple[str, int]:
    try:
        stat = path.stat()
    except FileNotFoundError as exc:
        raise FerpLabelRegistryError(f"FERP_LABELS_XLSX_NOT_FOUND: {path}") from exc
    if not path.is_file():
        raise FerpLabelRegistryError(f"FERP_LABELS_XLSX_NOT_FILE: {path}")
    return str(path.resolve()), int(stat.st_mtime_ns)


@lru_cache(maxsize=8)
def _load_registry_cached(path_text: str, mtime_ns: int) -> dict[str, dict[str, dict[str, str]]]:
    del mtime_ns
    if load_workbook is None:
        raise FerpLabelRegistryError("FERP_LABELS_OPENPYXL_NOT_INSTALLED")
    try:
        workbook = load_workbook(path_text, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - exercised through caller behavior.
        raise FerpLabelRegistryError(f"FERP_LABELS_XLSX_READ_FAILED: {path_text}: {exc}") from exc

    try:
        registry: dict[str, dict[str, FerpLabel]] = {}
        for sheet_name in workbook.sheetnames:
            sheet_registry = _extract_sheet_registry(workbook[sheet_name])
            for object_code, labels in sheet_registry.items():
                registry.setdefault(object_code, {}).update(labels)
    finally:
        with contextlib.suppress(Exception):
            workbook.close()

    missing_objects = sorted(SUPPORTED_OBJECTS - set(registry))
    if missing_objects:
        raise FerpLabelRegistryError(f"FERP_LABELS_OBJECTS_NOT_FOUND: {', '.join(missing_objects)}")

    return {
        object_code: {label_code: label.to_dict() for label_code, label in labels.items()}
        for object_code, labels in registry.items()
    }


def load_ferp_label_registry(path: str | Path | None = None) -> dict[str, dict[str, FerpLabel]]:
    source = Path(path) if path is not None else default_ferp_labels_path()
    path_text, mtime_ns = _validate_source_path(source)
    raw_registry = _load_registry_cached(path_text, mtime_ns)
    return {
        object_code: {
            label_code: FerpLabel(**label_data)
            for label_code, label_data in labels.items()
        }
        for object_code, labels in raw_registry.items()
    }


def get_labels_for_object(object_code: str, path: str | Path | None = None) -> dict[str, FerpLabel]:
    registry = load_ferp_label_registry(path)
    return dict(registry.get(str(object_code or "").strip().lower(), {}))


def find_label(object_code: str, label_code: str, path: str | Path | None = None) -> FerpLabel | None:
    labels = get_labels_for_object(object_code, path)
    return labels.get(str(label_code or "").strip())


def require_label(object_code: str, label_code: str, path: str | Path | None = None) -> FerpLabel:
    label = find_label(object_code, label_code, path)
    if label is None:
        raise FerpLabelRegistryError(f"FERP_LABEL_NOT_FOUND: {object_code}.{label_code}")
    return label


def validate_label_payload(
    object_code: str,
    payload: Any,
    path: str | Path | None = None,
) -> dict[str, Any]:
    normalized_object = str(object_code or "").strip().lower()
    label_payload = payload if isinstance(payload, dict) else {}
    warnings: list[str] = []
    known_labels: list[str] = []
    unknown_labels: list[str] = []
    missing_required_labels: list[str] = []

    if not normalized_object:
        warnings.append("FERP_OBJECT_REQUIRED")
        return {
            "valid": False,
            "known_labels": known_labels,
            "unknown_labels": sorted(str(key) for key in label_payload),
            "missing_required_labels": missing_required_labels,
            "warnings": warnings,
        }

    try:
        labels = get_labels_for_object(normalized_object, path)
    except FerpLabelRegistryError as exc:
        warnings.append(str(exc))
        return {
            "valid": False,
            "known_labels": known_labels,
            "unknown_labels": sorted(str(key) for key in label_payload),
            "missing_required_labels": missing_required_labels,
            "warnings": warnings,
        }

    if not labels:
        warnings.append(f"FERP_OBJECT_NOT_SUPPORTED: {normalized_object}")
        return {
            "valid": False,
            "known_labels": known_labels,
            "unknown_labels": sorted(str(key) for key in label_payload),
            "missing_required_labels": missing_required_labels,
            "warnings": warnings,
        }

    for label_code in sorted(str(key) for key in label_payload):
        if label_code in labels:
            known_labels.append(label_code)
        else:
            unknown_labels.append(label_code)

    required = REQUIRED_LABELS_BY_OBJECT.get(normalized_object, set())
    missing_required_labels = sorted(label_code for label_code in required if label_code not in label_payload)
    if unknown_labels:
        warnings.append(f"FERP_UNKNOWN_LABELS: {', '.join(unknown_labels)}")
    if missing_required_labels:
        warnings.append(f"FERP_MISSING_REQUIRED_LABELS: {', '.join(missing_required_labels)}")

    return {
        "valid": not unknown_labels and not missing_required_labels,
        "known_labels": known_labels,
        "unknown_labels": unknown_labels,
        "missing_required_labels": missing_required_labels,
        "warnings": warnings,
    }
