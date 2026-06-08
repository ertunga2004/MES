import argparse
import json
import os
import sys
from collections import defaultdict
from pathlib import Path
try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is not installed. Please install it to run this script.")
    sys.exit(1)


def parse_args():
    parser = argparse.ArgumentParser(description="Dry-run analyzer for mes.vision_events from Excel workbook")
    parser.add_argument("--workbook", type=str, required=True, help="Path to the Excel workbook (.xlsx)")
    return parser.parse_args()


def load_raw_payloads(wb):
    raw_logs_sheet = "99_Raw_Logs"
    payloads_by_item_event = {}
    raw_logs_count = 0
    
    if raw_logs_sheet not in wb.sheetnames:
        print(f"Warning: {raw_logs_sheet} sheet not found. Raw payload matching will be skipped.")
        return payloads_by_item_event, raw_logs_count

    ws = wb[raw_logs_sheet]
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c or "").strip().lower() for c in row]
            continue
            
        raw_logs_count += 1
        
        row_dict = dict(zip(headers, row))
        event_type = str(row_dict.get("event_type_code") or "").strip()
        item_id = str(row_dict.get("item_id") or "").strip()
        raw_payload = str(row_dict.get("raw_payload") or "").strip()
        
        if event_type == "vision_event" and item_id and raw_payload:
            payloads_by_item_event[item_id] = raw_payload
            
    return payloads_by_item_event, raw_logs_count


def main():
    args = parse_args()
    workbook_path = Path(args.workbook)
    
    if not workbook_path.is_file():
        print(f"Error: Workbook file not found at {workbook_path}")
        sys.exit(1)

    print(f"\nLoading workbook: {workbook_path}")
    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"Error: Could not parse workbook. {e}")
        sys.exit(1)

    vision_sheet = "6_Vision"
    if vision_sheet not in wb.sheetnames:
        print(f"Error: {vision_sheet} sheet not found in the workbook.")
        sys.exit(1)

    payloads_by_item, raw_logs_count = load_raw_payloads(wb)

    ws = wb[vision_sheet]
    
    headers = []
    candidates = []
    
    stats = {
        "workbook": workbook_path.name,
        "sheet_name": vision_sheet,
        "raw_logs_row_count": raw_logs_count,
        "raw_excel_data_row_count": 0,
        "skipped_blank_row_count": 0,
        "candidate_event_count": 0,
        "apply_safe_count": 0,
        "apply_unsafe_count": 0,
        "missing_event_key_count": 0,
        "missing_detected_at_count": 0,
        "missing_event_type_count": 0,
        "duplicate_external_ref_count": 0,
        "payload_missing_count": 0,
        "db_writes": False
    }

    seen_external_refs = set()
    duplicate_external_refs = set()
    item_id_counts = defaultdict(int)

    sample_safe = []
    sample_unsafe = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c or "").strip().lower() for c in row]
            continue
        
        stats["raw_excel_data_row_count"] += 1
        row_dict = dict(zip(headers, row))
        
        # Field extraction and normalization
        vision_track_id = str(row_dict.get("vision_track_id") or "").strip()
        event_key = str(row_dict.get("event_key") or "").strip()
        event_type = str(row_dict.get("event_type") or "").strip()
        item_id = str(row_dict.get("item_id") or "").strip()
        color_code = str(row_dict.get("color_code") or "").strip()
        classification = str(row_dict.get("classification") or row_dict.get("confidence_tier") or "").strip()
        vision_observed_at = str(row_dict.get("vision_observed_at") or "").strip()
        detected_at = str(row_dict.get("detected_at") or vision_observed_at or row_dict.get("event_time") or "").strip()
        correlation_status = str(row_dict.get("correlation_status") or "").strip()
        
        raw_payload = payloads_by_item.get(item_id, "") if item_id else ""
        
        # Blank row check
        if not any([vision_track_id, event_key, event_type, item_id, color_code, classification, vision_observed_at, detected_at, correlation_status, raw_payload]):
            stats["skipped_blank_row_count"] += 1
            continue

        # Natural key logic
        external_ref = ""
        is_fallback_key = False
        if event_key:
            external_ref = event_key
        elif vision_track_id:
            external_ref = vision_track_id
        else:
            if event_type and detected_at and item_id:
                external_ref = f"{event_type}_{detected_at}_{item_id}"
                is_fallback_key = True

        unsafe_reasons = []
        if not external_ref:
            unsafe_reasons.append("missing_stable_key")
            stats["missing_event_key_count"] += 1
        elif is_fallback_key:
            unsafe_reasons.append("fallback_key_used")
            
        if not event_type:
            unsafe_reasons.append("missing_event_type")
            stats["missing_event_type_count"] += 1
            
        if not detected_at:
            unsafe_reasons.append("missing_detected_at")
            stats["missing_detected_at_count"] += 1

        if not raw_payload:
            stats["payload_missing_count"] += 1
            # Warning only, doesn't make it unsafe

        if external_ref:
            if external_ref in seen_external_refs:
                duplicate_external_refs.add(external_ref)
                unsafe_reasons.append("duplicate_external_ref")
                stats["duplicate_external_ref_count"] += 1
            seen_external_refs.add(external_ref)
            
        if item_id:
            item_id_counts[item_id] += 1

        is_safe = len(unsafe_reasons) == 0
        stats["candidate_event_count"] += 1
        
        mapping = {
            "event_key": external_ref,
            "item_id": item_id,
            "event_type": event_type,
            "detected_at": detected_at,
            "source_system": "mes_web",
            "source_file": stats["workbook"],
            "external_ref": external_ref,
            "payload": raw_payload or "{}",
            "metadata": json.dumps({
                "color_code": color_code,
                "correlation_status": correlation_status,
                "classification": classification,
                "is_fallback_key": is_fallback_key
            }),
            "unsafe_reasons": unsafe_reasons
        }

        if is_safe:
            stats["apply_safe_count"] += 1
            if len(sample_safe) < 5:
                sample_safe.append(mapping)
        else:
            stats["apply_unsafe_count"] += 1
            if len(sample_unsafe) < 5:
                sample_unsafe.append(mapping)

    wb.close()

    print("\n" + "="*50)
    print(f" DRY-RUN RESULTS for {workbook_path.name}")
    print("="*50)
    for k, v in stats.items():
        print(f"  {k}: {v}")

    print("\n--- Items with multiple events (Top 5) ---")
    multiple_items = {k: v for k, v in item_id_counts.items() if v > 1}
    for i, (k, v) in enumerate(list(multiple_items.items())[:5]):
        print(f"  item_id: {k} -> {v} events")

    print("\n--- First 5 APPLY_SAFE Candidates ---")
    for i, s in enumerate(sample_safe, 1):
        print(f" {i}. external_ref={s['external_ref']}, type={s['event_type']}, item={s['item_id']}, payload_len={len(s['payload'])}")

    print("\n--- First 5 UNSAFE Candidates ---")
    for i, s in enumerate(sample_unsafe, 1):
        print(f" {i}. external_ref={s['external_ref']}, reasons={s['unsafe_reasons']}, type={s['event_type']}, item={s['item_id']}")

    print("==================================================")
    
    if stats["candidate_event_count"] == 0:
        print("Note: No vision event rows found in this workbook.")

    return 0

if __name__ == "__main__":
    sys.exit(main())
