from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is not installed. Please install it to run this script.", file=sys.stderr)
    sys.exit(1)

try:
    import dateutil.parser as dp
except ImportError:
    print("Error: python-dateutil is not installed. Please install it to run this script.", file=sys.stderr)
    sys.exit(1)

sys.dont_write_bytecode = True

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from mes_web.config import AppConfig
from mes_web.db.config import build_database_config

JsonObject = dict[str, Any]

UPSERT_SQL = """
INSERT INTO mes.vision_events (
    event_key,
    item_id,
    event_type,
    detected_at,
    source_system,
    source_file,
    external_ref,
    payload,
    metadata,
    created_at
) VALUES (
    %(event_key)s,
    %(item_id)s,
    %(event_type)s,
    %(detected_at)s,
    %(source_system)s,
    %(source_file)s,
    %(external_ref)s,
    %(payload)s,
    %(metadata)s,
    now()
)
"""

UPDATE_SQL = """
UPDATE mes.vision_events SET
    event_key = %(event_key)s,
    item_id = %(item_id)s,
    event_type = %(event_type)s,
    detected_at = %(detected_at)s,
    source_system = %(source_system)s,
    source_file = %(source_file)s,
    payload = %(payload)s,
    metadata = %(metadata)s
WHERE external_ref = %(external_ref)s
"""

def _resolve_path(value: str | Path) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return REPO_ROOT / path

def _import_psycopg() -> tuple[Any, Any]:
    try:
        import psycopg
        from psycopg.types.json import Jsonb
    except ModuleNotFoundError as exc:
        raise RuntimeError("psycopg is not installed") from exc
    return psycopg, Jsonb

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Mirror mes.vision_events from Excel workbook"
    )
    parser.add_argument(
        "--workbook",
        required=True,
        help="Path to the Excel workbook (.xlsx)",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Write to mes.vision_events. Requires MES_WEB_DB_ENABLED=true.",
    )
    return parser.parse_args()

def load_raw_payloads(wb):
    raw_logs_sheet = "99_Raw_Logs"
    payloads_by_item_event = {}
    
    if raw_logs_sheet not in wb.sheetnames:
        return payloads_by_item_event

    ws = wb[raw_logs_sheet]
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c or "").strip().lower() for c in row]
            continue
            
        row_dict = dict(zip(headers, row))
        event_type = str(row_dict.get("event_type_code") or "").strip()
        item_id = str(row_dict.get("item_id") or "").strip()
        raw_payload = str(row_dict.get("raw_payload") or "").strip()
        
        if event_type == "vision_event" and item_id and raw_payload:
            payloads_by_item_event[item_id] = raw_payload
            
    return payloads_by_item_event

def parse_detected_at(dt_str: str) -> tuple[datetime | None, str | None]:
    if not dt_str:
        return None, "empty string"
    try:
        dt = dp.parse(dt_str)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt, None
    except Exception as e:
        return None, str(e)

def main() -> int:
    args = parse_args()
    app_config = AppConfig.from_env()
    workbook_path = _resolve_path(args.workbook)

    if not workbook_path.is_file():
        print(f"Error: Workbook file not found at {workbook_path}", file=sys.stderr)
        return 1

    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"Error: Could not parse workbook. {e}", file=sys.stderr)
        return 1

    vision_sheet = "6_Vision"
    if vision_sheet not in wb.sheetnames:
        print(f"Error: {vision_sheet} sheet not found in the workbook.", file=sys.stderr)
        return 1

    payloads_by_item = load_raw_payloads(wb)
    ws = wb[vision_sheet]

    candidate_event_count = 0
    apply_safe_count = 0
    skipped_count = 0
    duplicate_external_ref_count = 0

    seen_external_refs = set()
    mapped_rows = []
    headers = []
    
    now_utc = datetime.now(timezone.utc)

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c or "").strip().lower() for c in row]
            continue
            
        row_dict = dict(zip(headers, row))
        
        vision_track_id = str(row_dict.get("vision_track_id") or "").strip()
        event_key = str(row_dict.get("event_key") or "").strip()
        event_type = str(row_dict.get("event_type") or "").strip()
        item_id = str(row_dict.get("item_id") or "").strip()
        color_code = str(row_dict.get("color_code") or "").strip()
        classification = str(row_dict.get("classification") or row_dict.get("confidence_tier") or "").strip()
        vision_observed_at = str(row_dict.get("vision_observed_at") or "").strip()
        detected_at_str = str(row_dict.get("detected_at") or vision_observed_at or row_dict.get("event_time") or "").strip()
        correlation_status = str(row_dict.get("correlation_status") or "").strip()
        
        raw_payload = payloads_by_item.get(item_id, "") if item_id else ""
        
        if not any([vision_track_id, event_key, event_type, item_id, color_code, classification, vision_observed_at, detected_at_str, correlation_status, raw_payload]):
            continue

        candidate_event_count += 1
        
        external_ref = ""
        if event_key:
            external_ref = event_key
        elif vision_track_id and event_type and detected_at_str:
            external_ref = f"{vision_track_id}_{event_type}_{detected_at_str}"

        unsafe_reasons = []
        if not external_ref:
            unsafe_reasons.append("missing_stable_key")
            
        if not event_type:
            unsafe_reasons.append("missing_event_type")

        dt_obj, parse_err = parse_detected_at(detected_at_str)
        if not dt_obj:
            unsafe_reasons.append(f"invalid_detected_at: {parse_err}")
        elif dt_obj > now_utc:
            unsafe_reasons.append("future_detected_at")
            print(f"Warning: Future timestamp detected {dt_obj} (now: {now_utc})")

        if external_ref:
            if external_ref in seen_external_refs:
                duplicate_external_ref_count += 1
                unsafe_reasons.append("duplicate_external_ref")
            else:
                seen_external_refs.add(external_ref)

        is_safe = len(unsafe_reasons) == 0
        
        try:
            payload_obj = json.loads(raw_payload) if raw_payload else row_dict
        except Exception:
            payload_obj = {"raw_fallback": raw_payload, "parsed_row": row_dict}
            
        if is_safe:
            apply_safe_count += 1
            mapped = {
                "event_key": event_key or None,
                "item_id": item_id or None,
                "event_type": event_type,
                "detected_at": dt_obj.isoformat(),
                "source_system": "mes_web",
                "source_file": workbook_path.name,
                "external_ref": external_ref,
                "payload": payload_obj,
                "metadata": {
                    "workbook": workbook_path.name,
                    "sheet_name": vision_sheet,
                    "excel_row_number": i + 1,
                    "vision_track_id": vision_track_id,
                    "color_code": color_code,
                    "classification": classification,
                    "correlation_status": correlation_status,
                    "dry_run_status": "APPLY_SAFE",
                    "mirrored_by": "scripts/mirror_vision_events_from_excel.py",
                    "mapped_at": now_utc.isoformat()
                }
            }
            mapped_rows.append(mapped)
        else:
            skipped_count += 1

    wb.close()

    print("--- Vision Events Mirror Summary ---")
    print(f"workbook: {workbook_path.name}")
    print(f"candidate_event_count: {candidate_event_count}")
    print(f"apply_safe_count: {apply_safe_count}")
    print(f"skipped_count: {skipped_count}")
    print(f"duplicate_external_ref_count: {duplicate_external_ref_count}")

    if not args.apply:
        print("db_writes: False")
        print("inserted: 0")
        print("updated: 0")
        return 0

    db_config = build_database_config(app_config)
    if not db_config.enabled:
        print("Error: --apply requires MES_WEB_DB_ENABLED=true; no DB connection was opened.", file=sys.stderr)
        return 1

    if not mapped_rows:
        print("No APPLY_SAFE vision events found; nothing to apply.")
        print("db_writes: False")
        return 0

    psycopg, Jsonb = _import_psycopg()
    existing_refs: set[str] = set()
    inserted = 0
    updated = 0

    with psycopg.connect(**db_config.connection_kwargs()) as connection:
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT external_ref FROM mes.vision_events")
                existing_refs = {str(row[0]) for row in cursor.fetchall()}
                for row in mapped_rows:
                    params = dict(row)
                    params["payload"] = Jsonb(row["payload"])
                    params["metadata"] = Jsonb(row["metadata"])
                    if row["external_ref"] in existing_refs:
                        cursor.execute(UPDATE_SQL, params)
                        updated += 1
                    else:
                        cursor.execute(UPSERT_SQL, params)
                        inserted += 1
                        existing_refs.add(row["external_ref"])
            connection.commit()
        except Exception as exc:
            connection.rollback()
            print(f"Vision event mirror apply failed: {type(exc).__name__}: {exc}", file=sys.stderr)
            return 1

    print("db_writes: True")
    print(f"inserted: {inserted}")
    print(f"updated: {updated}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
