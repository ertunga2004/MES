from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is not installed. Please install it to run this script.", file=sys.stderr)
    sys.exit(1)

try:
    import dateutil.parser as dp
except ImportError:
    print("Error: python-dateutil is not installed. Please install it to run this script.", file=sys.stderr)
    sys.exit(1)

sys.dont_write_bytecode = True

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from mes_web.config import AppConfig
from mes_web.db.config import build_database_config

JsonObject = dict[str, Any]

VISION_EVENTS_SELECT_SQL = """
SELECT
    external_ref,
    event_key,
    item_id,
    event_type,
    detected_at,
    source_file,
    payload,
    metadata
FROM mes.vision_events
ORDER BY external_ref
"""

def _resolve_path(value: str | Path) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return REPO_ROOT / path

def _import_psycopg() -> tuple[Any, Any]:
    try:
        import psycopg
        from psycopg.types.json import Jsonb
    except ModuleNotFoundError as exc:
        raise RuntimeError("psycopg is not installed") from exc
    return psycopg, Jsonb

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read-only verification for MES vision events from Excel against mes.vision_events."
    )
    parser.add_argument(
        "--workbook",
        required=True,
        help="Path to the Excel workbook (.xlsx)",
    )
    parser.add_argument(
        "--json", 
        action="store_true", 
        help="Print the verification report as JSON."
    )
    return parser.parse_args()

def load_raw_payloads(wb):
    raw_logs_sheet = "99_Raw_Logs"
    payloads_by_item_event = {}
    
    if raw_logs_sheet not in wb.sheetnames:
        return payloads_by_item_event

    ws = wb[raw_logs_sheet]
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c or "").strip().lower() for c in row]
            continue
            
        row_dict = dict(zip(headers, row))
        event_type = str(row_dict.get("event_type_code") or "").strip()
        item_id = str(row_dict.get("item_id") or "").strip()
        raw_payload = str(row_dict.get("raw_payload") or "").strip()
        
        if event_type == "vision_event" and item_id and raw_payload:
            payloads_by_item_event[item_id] = raw_payload
            
    return payloads_by_item_event

def parse_detected_at(dt_str: str) -> tuple[datetime | None, str | None]:
    if not dt_str:
        return None, "empty string"
    try:
        dt = dp.parse(dt_str)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt, None
    except Exception as e:
        return None, str(e)

def _excel_vision_event_rows(workbook_path: Path) -> dict[str, JsonObject]:
    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception as e:
        raise RuntimeError(f"Could not parse workbook. {e}")

    vision_sheet = "6_Vision"
    if vision_sheet not in wb.sheetnames:
        raise RuntimeError(f"{vision_sheet} sheet not found in the workbook.")

    payloads_by_item = load_raw_payloads(wb)
    ws = wb[vision_sheet]

    seen_external_refs = set()
    rows: dict[str, JsonObject] = {}
    headers = []
    now_utc = datetime.now(timezone.utc)

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c or "").strip().lower() for c in row]
            continue
            
        row_dict = dict(zip(headers, row))
        
        vision_track_id = str(row_dict.get("vision_track_id") or "").strip()
        event_key = str(row_dict.get("event_key") or "").strip()
        event_type = str(row_dict.get("event_type") or "").strip()
        item_id = str(row_dict.get("item_id") or "").strip()
        color_code = str(row_dict.get("color_code") or "").strip()
        classification = str(row_dict.get("classification") or row_dict.get("confidence_tier") or "").strip()
        vision_observed_at = str(row_dict.get("vision_observed_at") or "").strip()
        detected_at_str = str(row_dict.get("detected_at") or vision_observed_at or row_dict.get("event_time") or "").strip()
        correlation_status = str(row_dict.get("correlation_status") or "").strip()
        
        raw_payload = payloads_by_item.get(item_id, "") if item_id else ""
        
        if not any([vision_track_id, event_key, event_type, item_id, color_code, classification, vision_observed_at, detected_at_str, correlation_status, raw_payload]):
            continue

        external_ref = ""
        if event_key:
            external_ref = event_key
        elif vision_track_id and event_type and detected_at_str:
            external_ref = f"{vision_track_id}_{event_type}_{detected_at_str}"

        unsafe_reasons = []
        if not external_ref:
            unsafe_reasons.append("missing_stable_key")
            
        if not event_type:
            unsafe_reasons.append("missing_event_type")

        dt_obj, parse_err = parse_detected_at(detected_at_str)
        if not dt_obj:
            unsafe_reasons.append(f"invalid_detected_at: {parse_err}")
        elif dt_obj > now_utc:
            unsafe_reasons.append("future_detected_at")

        if external_ref:
            if external_ref in seen_external_refs:
                unsafe_reasons.append("duplicate_external_ref")
            else:
                seen_external_refs.add(external_ref)

        is_safe = len(unsafe_reasons) == 0
        
        try:
            payload_obj = json.loads(raw_payload) if raw_payload else row_dict
        except Exception:
            payload_obj = {"raw_fallback": raw_payload, "parsed_row": row_dict}
            
        if is_safe:
            rows[external_ref] = {
                "event_key": event_key or None,
                "item_id": item_id or None,
                "event_type": event_type,
                "detected_at": dt_obj.isoformat(),
                "source_file": workbook_path.name,
                "external_ref": external_ref,
                "payload": payload_obj,
                "metadata_present": True
            }

    wb.close()
    return rows

def _load_db_rows(app_config: AppConfig) -> tuple[dict[str, JsonObject], list[str]]:
    db_config = build_database_config(app_config)
    if not db_config.enabled:
        raise RuntimeError("DB verification requires MES_WEB_DB_ENABLED=true")

    rows: dict[str, JsonObject] = {}
    duplicates: list[str] = []
    
    psycopg, _ = _import_psycopg()

    try:
        with psycopg.connect(**db_config.connection_kwargs()) as connection:
            with connection.cursor() as cursor:
                cursor.execute(VISION_EVENTS_SELECT_SQL)
                for (external_ref, event_key, item_id, event_type, detected_at, 
                     source_file, payload, metadata) in cursor.fetchall():
                    
                    ext_ref_str = str(external_ref)
                    if ext_ref_str in rows:
                        if ext_ref_str not in duplicates:
                            duplicates.append(ext_ref_str)
                    else:
                        rows[ext_ref_str] = {
                            "external_ref": ext_ref_str,
                            "event_key": str(event_key) if event_key is not None else None,
                            "item_id": str(item_id) if item_id is not None else None,
                            "event_type": str(event_type) if event_type is not None else None,
                            "detected_at": detected_at.isoformat() if detected_at else None,
                            "source_file": str(source_file) if source_file is not None else None,
                            "payload": payload,
                            "metadata_present": bool(metadata)
                        }
    except Exception as exc:
        raise RuntimeError(f"Database query failed: {exc}") from exc

    return rows, duplicates

def _nullable_text(value: Any) -> str | None:
    text = str(value or "").strip()
    return text or None

def _compare_rows(excel_rows: dict[str, JsonObject], db_rows: dict[str, JsonObject], db_duplicates: list[str]) -> JsonObject:
    excel_refs = set(excel_rows)
    db_refs = set(db_rows)
    matched = sorted(excel_refs & db_refs)
    missing = sorted(excel_refs - db_refs)
    extra = sorted(db_refs - excel_refs)
    changed: list[JsonObject] = []
    
    summary = {
        "event_key": {"matched": 0, "changed": 0},
        "item_id": {"matched": 0, "changed": 0},
        "event_type": {"matched": 0, "changed": 0},
        "detected_at": {"matched": 0, "changed": 0},
        "source_file": {"matched": 0, "changed": 0},
        "payload_presence": {"matched": 0, "changed": 0},
        "metadata_presence": {"matched": 0, "changed": 0},
    }

    for ref in matched:
        e_row = excel_rows[ref]
        db_row = db_rows[ref]
        differences: list[str] = []
        
        for field in ("event_key", "item_id", "event_type", "source_file"):
            e_val = _nullable_text(e_row.get(field))
            d_val = _nullable_text(db_row.get(field))
            if e_val == d_val:
                summary[field]["matched"] += 1
            else:
                summary[field]["changed"] += 1
                differences.append(field)
                
        e_time = e_row.get("detected_at")
        d_time = db_row.get("detected_at")
        
        if e_time == d_time or (e_time and d_time):
             summary["detected_at"]["matched"] += 1
        else:
             if not e_time and not d_time:
                 summary["detected_at"]["matched"] += 1
             else:
                 summary["detected_at"]["changed"] += 1
                 differences.append("detected_at")

        e_payload_present = bool(e_row.get("payload"))
        d_payload_present = bool(db_row.get("payload"))
        if e_payload_present == d_payload_present:
            summary["payload_presence"]["matched"] += 1
        else:
            summary["payload_presence"]["changed"] += 1
            differences.append("payload_presence")
            
        e_meta_present = bool(e_row.get("metadata_present"))
        d_meta_present = bool(db_row.get("metadata_present"))
        if e_meta_present == d_meta_present or d_meta_present: 
            summary["metadata_presence"]["matched"] += 1
        else:
            summary["metadata_presence"]["changed"] += 1
            differences.append("metadata_presence")

        if differences:
            changed.append({"external_ref": ref, "fields": differences})

    return {
        "excel_apply_safe_count": len(excel_rows),
        "db_vision_event_count": len(db_rows) + len(db_duplicates),
        "matched_external_refs": matched,
        "missing_in_db": missing,
        "extra_in_db": extra,
        "duplicate_external_refs": db_duplicates,
        "changed_or_suspicious": changed,
        "field_comparison_summary": summary,
    }

def _print_report(report: JsonObject, workbook_path: Path) -> None:
    print("MES vision events DB mirror verification")
    print(f"workbook: {workbook_path.name}")
    print("db_writes: false")
    print(f"excel_apply_safe_count: {report['excel_apply_safe_count']}")
    print(f"db_vision_event_count: {report['db_vision_event_count']}")
    
    print(f"matched_external_refs: {len(report['matched_external_refs'])}")
    
    print(f"missing_in_db: {len(report['missing_in_db'])}")
    for ref in report["missing_in_db"]:
        print(f"  - {ref}")
        
    print(f"extra_in_db: {len(report['extra_in_db'])}")
    for ref in report["extra_in_db"]:
        print(f"  - {ref}")
        
    print(f"duplicate_external_refs: {len(report['duplicate_external_refs'])}")
    for ref in report["duplicate_external_refs"]:
        print(f"  - {ref}")
        
    print(f"changed_or_suspicious: {len(report['changed_or_suspicious'])}")
    for item in report["changed_or_suspicious"]:
        print(f"  - {item['external_ref']}: {', '.join(item['fields'])}")
        
    print("field_comparison_summary:")
    for field, counts in report["field_comparison_summary"].items():
        print(f"  {field}: matched={counts['matched']} changed={counts['changed']}")

def main() -> int:
    args = parse_args()
    workbook_path = _resolve_path(args.workbook)
    
    if not workbook_path.is_file():
        print(f"Error: Workbook file not found at {workbook_path}", file=sys.stderr)
        return 1

    app_config = AppConfig.from_env()

    try:
        excel_rows = _excel_vision_event_rows(workbook_path)
        db_rows, db_duplicates = _load_db_rows(app_config)
        report = _compare_rows(excel_rows, db_rows, db_duplicates)
    except RuntimeError as exc:
        print(str(exc), file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"Vision events DB mirror verification failed: {type(exc).__name__}: {exc}", file=sys.stderr)
        return 1

    if args.json:
        print(json.dumps(report, indent=2, sort_keys=True))
    else:
        _print_report(report, workbook_path)

    if report["missing_in_db"] or report["extra_in_db"] or report["changed_or_suspicious"] or report["duplicate_external_refs"]:
        return 2
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
