from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from mes_web.excel_runtime import INVENTORY_SHEET_NAME, MAINTENANCE_SHEET_NAME, RAW_LOG_SHEET_NAME, WORK_ORDER_SHEET_NAME, WorkbookProjector


class WorkbookProjectorTests(unittest.TestCase):
    def setUp(self) -> None:
        self.projector = WorkbookProjector()

    def test_measurement_queue_and_completion_flow_into_rows(self) -> None:
        measurement_rows = self.projector.consume_mega_log(
            "MEGA|TCS3200|STATE=MEASURING|ITEM_ID=42|MEASURE_ID=8|FINAL=KIRMIZI|FINAL_SOURCE=CORE_STABLE|SEARCH_HINT=SARI|SEARCH_HINT_WIN=2|SEARCH_HINT_SECOND=1|SEARCH_HINT_STRONG=1|SEARCH_HINT_FALLBACK_ALLOWED=1|CORE_USED=1|CORE_N=6|OBJ_N=10|MEDIAN_NEAREST=SARI|SCORE_NEAREST=SARI|MED_R=10|MED_G=20|MED_B=30|MED_D_R=11|MED_D_Y=22|MED_D_B=33|MED_D_X=44|X_R=101|X_G=102|X_B=103|MED_OBJ=1|CONF=0|CORE_STR_MIN=12|CORE_STR_MAX=24|VOTE_WIN=6|VOTE_SECOND=1|VOTE_CLASSIFIED=6|VOTE_BOS=3|VOTE_R=0|VOTE_Y=1|VOTE_B=2|VOTE_CAL=0|TOT_R=4|TOT_Y=5|TOT_B=6|TOT_BOS=7|TOT_CAL=0",
            "2026-04-02T10:15:25Z",
        )
        queue_rows = self.projector.consume_mega_log(
            "MEGA|AUTO|QUEUE=ENQ|ITEM_ID=42|MEASURE_ID=8|COLOR=KIRMIZI|DECISION_SOURCE=CORE_STABLE|TRAVEL_MS=640",
            "2026-04-02T10:15:26Z",
        )
        released_rows = self.projector.consume_mega_log(
            "MEGA|ROBOT|EVENT=RELEASED|ITEM_ID=42|MEASURE_ID=8|TRIGGER=TIMER",
            "2026-04-02T10:15:29Z",
        )
        completed_rows = self.projector.consume_mega_log(
            "MEGA|AUTO|STATE=WAIT_ARM|EVENT=PICKPLACE_DONE|ITEM_ID=42|MEASURE_ID=8|COLOR=KIRMIZI|DECISION_SOURCE=CORE_STABLE|TRIGGER=TIMER|PENDING=0",
            "2026-04-02T10:15:29.300Z",
        )

        self.assertIn("2_Olcumler", measurement_rows)
        self.assertEqual(measurement_rows["2_Olcumler"][0]["item_id"], "42")
        self.assertEqual(measurement_rows["2_Olcumler"][0]["final_color_code"], "red")
        self.assertEqual(measurement_rows["2_Olcumler"][0]["search_hint"], "SARI")
        self.assertEqual(measurement_rows["2_Olcumler"][0]["search_hint_win"], 2)
        self.assertEqual(measurement_rows["2_Olcumler"][0]["core_n"], 6)
        self.assertEqual(measurement_rows["2_Olcumler"][0]["med_d_x"], 44)
        self.assertEqual(measurement_rows["2_Olcumler"][0]["x_b"], 103)
        self.assertEqual(measurement_rows["2_Olcumler"][0]["vote_x"], 3)
        self.assertEqual(measurement_rows["2_Olcumler"][0]["tot_x"], 7)
        self.assertEqual(measurement_rows["2_Olcumler"][0]["measurement_error_flag"], 1)
        self.assertEqual(measurement_rows["2_Olcumler"][0]["measurement_error_reason"], "confidence=0")
        self.assertEqual(queue_rows["1_Olay_Logu"][0]["event_type_code"], "queue_enq")
        self.assertEqual(queue_rows["1_Olay_Logu"][0]["mega_state_id"], 7)
        self.assertNotIn("4_Uretim_Tamamlanan", released_rows)
        self.assertIn("4_Uretim_Tamamlanan", completed_rows)
        self.assertEqual(completed_rows["4_Uretim_Tamamlanan"][0]["item_id"], "42")
        self.assertEqual(completed_rows["4_Uretim_Tamamlanan"][0]["status_code"], "COMPLETED")
        self.assertEqual(completed_rows["4_Uretim_Tamamlanan"][0]["travel_ms"], 640)
        self.assertEqual(completed_rows["4_Uretim_Tamamlanan"][0]["decision_source_code"], "CORE_STABLE")
        self.assertEqual(completed_rows["4_Uretim_Tamamlanan"][0]["detected_at"], "2026-04-02T10:15:25Z")
        self.assertEqual(completed_rows["4_Uretim_Tamamlanan"][0]["completed_at"], "2026-04-02T10:15:29Z")
        self.assertEqual(completed_rows["4_Uretim_Tamamlanan"][0]["flow_ms"], 4000)
        self.assertEqual(completed_rows["4_Uretim_Tamamlanan"][0]["cycle_ms"], "")

    def test_vision_event_creates_vision_and_raw_rows(self) -> None:
        rows = self.projector.consume_vision_event(
            {
                "event": "line_crossed",
                "color_name": "blue",
                "track_id": 17,
                "frame_index": 33,
                "bbox": {"x1": 1, "y1": 2, "x2": 3, "y2": 4},
            },
            "2026-04-02T10:15:27Z",
        )

        self.assertIn("6_Vision", rows)
        self.assertEqual(rows["6_Vision"][0]["vision_track_id"], "17")
        self.assertEqual(rows["6_Vision"][0]["color_code"], "blue")
        self.assertEqual(rows[RAW_LOG_SHEET_NAME][0]["event_type_code"], "vision_event")

    def test_local_counts_reset_creates_system_rows(self) -> None:
        rows = self.projector.consume_local_counts_reset("2026-04-02T10:15:30Z")

        self.assertEqual(rows["1_Olay_Logu"][0]["source_code"], "system")
        self.assertEqual(rows["1_Olay_Logu"][0]["raw_line"], "SYSTEM|COUNTS|RESET")
        self.assertEqual(rows[RAW_LOG_SHEET_NAME][0]["source_topic"], "local/system")

    def test_tablet_oee_log_creates_workbook_rows(self) -> None:
        rows = self.projector.consume_tablet_log(
            "|Tablet|OEE|OEE:74.5|KULL:80.0|PERF:90.0|KALITE:95.0|KIRMIZI_S:1|KIRMIZI_R:0|KIRMIZI_H:0|SARI_S:2|SARI_R:1|SARI_H:0|MAVI_S:3|MAVI_R:0|MAVI_H:1",
            "2026-04-02T10:15:30Z",
        )

        self.assertEqual(rows[RAW_LOG_SHEET_NAME][0]["source_topic"], "sau/iot/mega/konveyor/tablet/log")
        self.assertEqual(rows[RAW_LOG_SHEET_NAME][0]["event_type_code"], "tablet_oee_snapshot")
        self.assertEqual(rows["1_Olay_Logu"][0]["source_code"], "tablet")
        self.assertEqual(rows["1_Olay_Logu"][0]["event_type_code"], "tablet_oee_snapshot")
        self.assertEqual(rows["1_Olay_Logu"][0]["oee_snapshot_id"], 1)
        self.assertIn("5_OEE_Anliklari", rows)
        self.assertEqual(rows["5_OEE_Anliklari"][0]["event_log_id"], rows["1_Olay_Logu"][0]["log_event_id"])
        self.assertEqual(rows["5_OEE_Anliklari"][0]["oee"], 74.5)
        self.assertEqual(rows["5_OEE_Anliklari"][0]["mavi_h"], 1)
        self.assertIn("toplam=8", rows["1_Olay_Logu"][0]["notes"])

    def test_tablet_fault_log_creates_workbook_rows(self) -> None:
        rows = self.projector.consume_tablet_log(
            "|Tablet|Ariza|DURUM:BASLADI|NEDEN:Motor Koruma|SURE_DK:4.5",
            "2026-04-02T10:15:31Z",
        )

        self.assertEqual(rows[RAW_LOG_SHEET_NAME][0]["event_type_code"], "tablet_fault")
        self.assertEqual(rows["1_Olay_Logu"][0]["event_type_code"], "tablet_fault")
        self.assertIn("neden=Motor Koruma", rows["1_Olay_Logu"][0]["notes"])
        self.assertIn("sure_ms=270000", rows["1_Olay_Logu"][0]["notes"])

    def test_system_oee_control_logs_create_workbook_rows(self) -> None:
        rows = self.projector.consume_system_oee_log(
            "SYSTEM|OEE|SET_TARGET_QTY|24",
            "2026-04-02T10:15:32Z",
        )

        self.assertEqual(rows[RAW_LOG_SHEET_NAME][0]["source_topic"], "local/oee")
        self.assertEqual(rows[RAW_LOG_SHEET_NAME][0]["event_type_code"], "oee_control")
        self.assertEqual(rows["1_Olay_Logu"][0]["event_summary_tr"], "Hedef guncellendi: 24")

    def test_shift_start_system_log_creates_workbook_rows(self) -> None:
        rows = self.projector.consume_system_oee_log(
            "|Tablet|Sistem| OLAY:VARDIYA_BASLADI|VARDIYA:SHIFT-A|PLAN_BASLANGIC:02.04.2026 08:00:00|PLAN_BITIS:02.04.2026 16:00:00|PERF_MOD:IDEAL_CYCLE|HEDEF:24|IDEAL_CYCLE_SN:1.8|PLANLI_DURUS_DK:15.0",
            "2026-04-02T10:15:33Z",
        )

        self.assertEqual(rows[RAW_LOG_SHEET_NAME][0]["event_type_code"], "shift_start")
        self.assertEqual(rows["1_Olay_Logu"][0]["event_type_code"], "shift_start")
        self.assertIn("planned_stop_ms=900000", rows["1_Olay_Logu"][0]["notes"])
        self.assertIn("planned_stop_dk=15.0", rows["1_Olay_Logu"][0]["notes"])

    def test_pickplace_return_done_creates_parsed_event_row(self) -> None:
        rows = self.projector.consume_mega_log(
            "MEGA|AUTO|STATE=SEARCHING|EVENT=PICKPLACE_RETURN_DONE|ITEM_ID=42|MEASURE_ID=8|COLOR=MAVI|DECISION_SOURCE=CORE_STABLE|TRIGGER=EARLY|PENDING=0",
            "2026-04-02T10:15:31Z",
        )

        self.assertEqual(rows[RAW_LOG_SHEET_NAME][0]["parsed_flag"], 1)
        self.assertEqual(rows[RAW_LOG_SHEET_NAME][0]["event_type_code"], "pickplace_return_done")
        self.assertEqual(rows["1_Olay_Logu"][0]["event_type_code"], "pickplace_return_done")
        self.assertEqual(rows["1_Olay_Logu"][0]["event_type_id"], 10)
        self.assertEqual(rows["1_Olay_Logu"][0]["station_id"], 3)

    def test_pick_drop_reached_creates_parsed_event_row(self) -> None:
        rows = self.projector.consume_mega_log(
            "MEGA|ROBOT|PICKPLACE=DROP_REACHED|ITEM_ID=42|MEASURE_ID=8|TRIGGER=TIMER|LIMIT=LIM23",
            "2026-04-02T10:15:28Z",
        )

        self.assertEqual(rows[RAW_LOG_SHEET_NAME][0]["parsed_flag"], 1)
        self.assertEqual(rows[RAW_LOG_SHEET_NAME][0]["event_type_code"], "pick_drop_reached")
        self.assertEqual(rows["1_Olay_Logu"][0]["event_type_code"], "pick_drop_reached")
        self.assertEqual(rows["1_Olay_Logu"][0]["event_type_id"], 20)
        self.assertEqual(rows["1_Olay_Logu"][0]["station_id"], 3)

    def test_pick_drop_reached_uses_active_pick_when_firmware_omits_identifiers(self) -> None:
        self.projector.consume_mega_log(
            "MEGA|AUTO|QUEUE=ENQ|ITEM_ID=42|MEASURE_ID=8|COLOR=MAVI|DECISION_SOURCE=CORE_STABLE|TRAVEL_MS=640",
            "2026-04-02T10:15:26Z",
        )
        self.projector.consume_mega_log(
            "MEGA|AUTO|STATE=WAIT_ARM|EVENT=ARM_POSITION_REACHED|ITEM_ID=42|MEASURE_ID=8|COLOR=MAVI|DECISION_SOURCE=CORE_STABLE|REVIEW=0|TRIGGER=TIMER",
            "2026-04-02T10:15:27Z",
        )

        rows = self.projector.consume_mega_log(
            "MEGA|ROBOT|PICKPLACE=DROP_REACHED|LIMIT=LIM23|TRIGGER=TIMER",
            "2026-04-02T10:15:28Z",
        )

        self.assertEqual(rows["1_Olay_Logu"][0]["item_id"], "42")
        self.assertEqual(rows["1_Olay_Logu"][0]["measure_id"], "8")
        self.assertEqual(rows[RAW_LOG_SHEET_NAME][0]["item_id"], "42")
        self.assertEqual(rows[RAW_LOG_SHEET_NAME][0]["measure_id"], "8")

    def test_vision_correction_and_early_pick_flow_reaches_completed_row(self) -> None:
        self.projector.consume_mega_log(
            "MEGA|AUTO|QUEUE=ENQ|ITEM_ID=42|MEASURE_ID=8|COLOR=KIRMIZI|DECISION_SOURCE=CORE_STABLE|REVIEW=0|TRAVEL_MS=640",
            "2026-04-02T10:15:26Z",
        )
        self.projector.consume_vision_event(
            {
                "event": "line_crossed",
                "item_id": "42",
                "measure_id": "8",
                "color_name": "blue",
                "track_id": 17,
                "confidence": 0.91,
                "confidence_tier": "high",
                "correlation_status": "MATCHED",
                "decision_applied": True,
                "review_required": False,
                "observed_at": "2026-04-02T10:15:27Z",
                "published_at": "2026-04-02T10:15:27.040Z",
            },
            "2026-04-02T10:15:27.080Z",
        )
        early_rows = self.projector.consume_early_pick_request("42", "2026-04-02T10:15:27.100Z")
        self.projector.consume_mega_log(
            "MEGA|AUTO|STATE=WAIT_ARM|EVENT=ARM_POSITION_REACHED|ITEM_ID=42|MEASURE_ID=8|COLOR=MAVI|DECISION_SOURCE=CORE_STABLE|REVIEW=0|TRIGGER=EARLY",
            "2026-04-02T10:15:27.150Z",
        )
        self.projector.consume_mega_log(
            "MEGA|ROBOT|EVENT=RELEASED|ITEM_ID=42|MEASURE_ID=8|TRIGGER=EARLY",
            "2026-04-02T10:15:29Z",
        )
        completed_rows = self.projector.consume_mega_log(
            "MEGA|AUTO|STATE=WAIT_ARM|EVENT=PICKPLACE_DONE|ITEM_ID=42|MEASURE_ID=8|COLOR=MAVI|DECISION_SOURCE=CORE_STABLE|REVIEW=0|TRIGGER=EARLY|PENDING=0",
            "2026-04-02T10:15:29.300Z",
        )

        self.assertEqual(early_rows["1_Olay_Logu"][0]["event_type_code"], "early_pick_request")
        row = completed_rows["4_Uretim_Tamamlanan"][0]
        self.assertEqual(row["sensor_color_code"], "red")
        self.assertEqual(row["vision_color_code"], "blue")
        self.assertEqual(row["final_color_code"], "blue")
        self.assertEqual(row["decision_source_code"], "VISION")
        self.assertEqual(row["finalization_reason"], "VISION_CORRECTED_MISMATCH")
        self.assertEqual(row["correlation_status"], "matched")
        self.assertEqual(row["mismatch_flag"], 1)
        self.assertEqual(row["pick_trigger_source"], "EARLY")
        self.assertEqual(row["early_pick_triggered"], 1)
        self.assertEqual(row["early_pick_request_sent_at"], "2026-04-02T10:15:27.100Z")
        self.assertEqual(row["early_pick_accepted_at"], "2026-04-02T10:15:27.150Z")
        self.assertEqual(row["flow_ms"], 3000)

    def test_late_vision_event_marks_audit_columns(self) -> None:
        rows = self.projector.consume_vision_event(
            {
                "event": "line_crossed",
                "item_id": "42",
                "measure_id": "8",
                "color_name": "yellow",
                "track_id": 21,
                "confidence": 0.95,
                "confidence_tier": "high",
                "correlation_status": "LATE",
                "late_vision_audit_flag": True,
                "decision_applied": False,
                "review_required": True,
                "observed_at": "2026-04-02T10:15:27Z",
                "published_at": "2026-04-02T10:15:27.040Z",
            },
            "2026-04-02T10:15:27.500Z",
        )

        row = rows["6_Vision"][0]
        self.assertEqual(row["correlation_status"], "late")
        self.assertEqual(row["late_vision_audit_flag"], 1)
        self.assertEqual(row["vision_observed_at"], "2026-04-02T10:15:27Z")
        self.assertEqual(row["vision_published_at"], "2026-04-02T10:15:27.040Z")
        self.assertEqual(row["vision_received_at"], "2026-04-02T10:15:27.500Z")

    def test_late_vision_event_updates_completed_row_after_completion(self) -> None:
        self.projector.consume_mega_log(
            "MEGA|AUTO|QUEUE=ENQ|ITEM_ID=42|MEASURE_ID=8|COLOR=SARI|DECISION_SOURCE=CORE_STABLE|REVIEW=0|TRAVEL_MS=640",
            "2026-04-02T10:15:26Z",
        )
        self.projector.consume_mega_log(
            "MEGA|ROBOT|EVENT=RELEASED|ITEM_ID=42|MEASURE_ID=8|TRIGGER=TIMER",
            "2026-04-02T10:15:29Z",
        )
        completed_rows = self.projector.consume_mega_log(
            "MEGA|AUTO|STATE=WAIT_ARM|EVENT=PICKPLACE_DONE|ITEM_ID=42|MEASURE_ID=8|COLOR=SARI|DECISION_SOURCE=CORE_STABLE|REVIEW=0|TRIGGER=TIMER|PENDING=0",
            "2026-04-02T10:15:29.300Z",
        )

        self.projector.consume_vision_event(
            {
                "event": "line_crossed",
                "item_id": "42",
                "measure_id": "8",
                "color_name": "red",
                "track_id": 22,
                "confidence": 0.96,
                "confidence_tier": "high",
                "correlation_status": "LATE",
                "late_vision_audit_flag": True,
                "decision_applied": False,
                "review_required": True,
            },
            "2026-04-02T10:15:30Z",
        )

        row = completed_rows["4_Uretim_Tamamlanan"][0]
        self.assertEqual(row["vision_color_code"], "red")
        self.assertEqual(row["correlation_status"], "late")
        self.assertEqual(row["review_required"], 1)
        self.assertEqual(row["status_code"], "COMPLETED_REVIEW")
        self.assertEqual(row["finalization_reason"], "SENSOR_LATE_VISION")
        self.assertEqual(self.projector.pending_completed_row_update, row)

    def test_quality_override_updates_completed_row_final_quality(self) -> None:
        self.projector.consume_mega_log(
            "MEGA|AUTO|QUEUE=ENQ|ITEM_ID=42|MEASURE_ID=8|COLOR=KIRMIZI|DECISION_SOURCE=CORE_STABLE|TRAVEL_MS=640",
            "2026-04-02T10:15:26Z",
        )
        self.projector.consume_mega_log(
            "MEGA|ROBOT|EVENT=RELEASED|ITEM_ID=42|MEASURE_ID=8|TRIGGER=TIMER",
            "2026-04-02T10:15:29Z",
        )
        completed_rows = self.projector.consume_mega_log(
            "MEGA|AUTO|STATE=WAIT_ARM|EVENT=PICKPLACE_DONE|ITEM_ID=42|MEASURE_ID=8|COLOR=KIRMIZI|DECISION_SOURCE=CORE_STABLE|TRIGGER=TIMER|PENDING=0",
            "2026-04-02T10:15:29.300Z",
        )

        self.projector.apply_quality_override("42", "SCRAP", "2026-04-02T10:18:00Z")

        row = completed_rows["4_Uretim_Tamamlanan"][0]
        self.assertEqual(row["status_code"], "COMPLETED_SCRAP")
        self.assertEqual(row["final_quality_code"], "SCRAP")
        self.assertEqual(row["override_flag"], 1)
        self.assertEqual(row["override_source_code"], "MANUAL")

    def test_quality_override_writes_back_to_workbook_row(self) -> None:
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl is not installed")
        from mes_web.excel_runtime import COMPLETED_COLUMNS, ExcelRuntimeSink

        self.projector.consume_mega_log(
            "MEGA|AUTO|QUEUE=ENQ|ITEM_ID=42|MEASURE_ID=8|COLOR=KIRMIZI|DECISION_SOURCE=CORE_STABLE|TRAVEL_MS=640",
            "2026-04-02T10:15:26+03:00",
        )
        self.projector.consume_mega_log(
            "MEGA|ROBOT|EVENT=RELEASED|ITEM_ID=42|MEASURE_ID=8|TRIGGER=TIMER",
            "2026-04-02T10:15:29+03:00",
        )
        completed_rows = self.projector.consume_mega_log(
            "MEGA|AUTO|STATE=WAIT_ARM|EVENT=PICKPLACE_DONE|ITEM_ID=42|MEASURE_ID=8|COLOR=KIRMIZI|DECISION_SOURCE=CORE_STABLE|TRIGGER=TIMER|PENDING=0",
            "2026-04-02T10:15:29.300+03:00",
        )
        row = self.projector.apply_quality_override("42", "SCRAP", "2026-04-02T10:18:00+03:00")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "4_Uretim_Tamamlanan"
        for col_index, header in enumerate(COMPLETED_COLUMNS, start=1):
            sheet.cell(1, col_index, header)
            sheet.cell(2, col_index, completed_rows["4_Uretim_Tamamlanan"][0].get(header, ""))

        sink = ExcelRuntimeSink.__new__(ExcelRuntimeSink)
        sink._update_completed_sheet_row(sheet, row)

        self.assertEqual(sheet.cell(2, COMPLETED_COLUMNS.index("status_code") + 1).value, "COMPLETED_SCRAP")
        self.assertEqual(sheet.cell(2, COMPLETED_COLUMNS.index("final_quality_code") + 1).value, "SCRAP")
        self.assertEqual(sheet.cell(2, COMPLETED_COLUMNS.index("override_applied_at") + 1).value, "2026-04-02T10:18:00+03:00")

    def test_sink_strips_illegal_excel_characters_before_row_write(self) -> None:
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl is not installed")
        from mes_web.excel_runtime import ExcelRuntimeSink, RAW_LOG_COLUMNS

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = RAW_LOG_SHEET_NAME
        for col_index, header in enumerate(RAW_LOG_COLUMNS, start=1):
            sheet.cell(1, col_index, header)

        sink = ExcelRuntimeSink.__new__(ExcelRuntimeSink)
        sink._write_sheet_row(
            sheet,
            2,
            RAW_LOG_COLUMNS,
            {
                "raw_log_id": 1,
                "source_code": "mega",
                "event_type_code": "raw",
                "notes": "bad\x00note\x0btext",
                "raw_payload": "xx\x01RL$q\x1fyy",
            },
        )

        self.assertEqual(sheet.cell(2, RAW_LOG_COLUMNS.index("notes") + 1).value, "badnotetext")
        self.assertEqual(sheet.cell(2, RAW_LOG_COLUMNS.index("raw_payload") + 1).value, "xxRL$qyy")

    def test_cycle_ms_is_gap_between_two_release_events(self) -> None:
        self.projector.consume_mega_log(
            "MEGA|TCS3200|STATE=MEASURING|ITEM_ID=42|MEASURE_ID=8|FINAL=KIRMIZI|FINAL_SOURCE=CORE_STABLE",
            "2026-04-02T10:15:25Z",
        )
        self.projector.consume_mega_log(
            "MEGA|AUTO|QUEUE=ENQ|ITEM_ID=42|MEASURE_ID=8|COLOR=KIRMIZI|DECISION_SOURCE=CORE_STABLE|TRAVEL_MS=640",
            "2026-04-02T10:15:26Z",
        )
        self.projector.consume_mega_log(
            "MEGA|ROBOT|EVENT=RELEASED|ITEM_ID=42|MEASURE_ID=8|TRIGGER=TIMER",
            "2026-04-02T10:15:29Z",
        )
        first_rows = self.projector.consume_mega_log(
            "MEGA|AUTO|STATE=WAIT_ARM|EVENT=PICKPLACE_DONE|ITEM_ID=42|MEASURE_ID=8|COLOR=KIRMIZI|DECISION_SOURCE=CORE_STABLE|TRIGGER=TIMER|PENDING=0",
            "2026-04-02T10:15:29.300Z",
        )

        self.projector.consume_mega_log(
            "MEGA|TCS3200|STATE=MEASURING|ITEM_ID=43|MEASURE_ID=9|FINAL=MAVI|FINAL_SOURCE=CORE_STABLE",
            "2026-04-02T10:15:30Z",
        )
        self.projector.consume_mega_log(
            "MEGA|AUTO|QUEUE=ENQ|ITEM_ID=43|MEASURE_ID=9|COLOR=MAVI|DECISION_SOURCE=CORE_STABLE|TRAVEL_MS=640",
            "2026-04-02T10:15:31Z",
        )
        self.projector.consume_mega_log(
            "MEGA|ROBOT|EVENT=RELEASED|ITEM_ID=43|MEASURE_ID=9|TRIGGER=TIMER",
            "2026-04-02T10:15:35Z",
        )
        second_rows = self.projector.consume_mega_log(
            "MEGA|AUTO|STATE=WAIT_ARM|EVENT=PICKPLACE_DONE|ITEM_ID=43|MEASURE_ID=9|COLOR=MAVI|DECISION_SOURCE=CORE_STABLE|TRIGGER=TIMER|PENDING=0",
            "2026-04-02T10:15:35.300Z",
        )

        self.assertEqual(first_rows["4_Uretim_Tamamlanan"][0]["cycle_ms"], "")
        self.assertEqual(second_rows["4_Uretim_Tamamlanan"][0]["flow_ms"], 5000)
        self.assertEqual(second_rows["4_Uretim_Tamamlanan"][0]["cycle_ms"], 6000)

    def test_work_order_sheet_writes_availability_and_oee_percentages(self) -> None:
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl is not installed")
        from mes_web.excel_runtime import ExcelRuntimeSink, INVENTORY_COLUMNS, WORK_ORDER_COLUMNS

        workbook = Workbook()
        work_order_sheet = workbook.active
        work_order_sheet.title = WORK_ORDER_SHEET_NAME
        inventory_sheet = workbook.create_sheet(INVENTORY_SHEET_NAME)
        sink = ExcelRuntimeSink.__new__(ExcelRuntimeSink)
        sink._ensure_sheet_layout(work_order_sheet, WORK_ORDER_COLUMNS)
        sink._ensure_sheet_layout(inventory_sheet, INVENTORY_COLUMNS)

        sink._sync_work_order_sheets(
            workbook,
            {
                "itemsById": {
                    "1": {"work_order_id": "WO-1", "completed_at": "2026-04-02T08:00:30+03:00", "classification": "GOOD"},
                    "2": {"work_order_id": "WO-1", "completed_at": "2026-04-02T08:00:55+03:00", "classification": "GOOD"},
                },
                "faultHistory": [
                    {"startedAt": "2026-04-02T08:00:10+03:00", "endedAt": "2026-04-02T08:00:20+03:00"},
                ],
                "workOrders": {
                    "ordersById": {
                        "WO-1": {
                            "orderId": "WO-1",
                            "erpType": "Is Emirleri",
                            "stockCode": "BOX-RED",
                            "stockName": "Kirmizi Kutu",
                            "productColor": "red",
                            "quantity": 2,
                            "completedQty": 2,
                            "productionQty": 2,
                            "inventoryConsumedQty": 0,
                            "startedAt": "2026-04-02T08:00:00+03:00",
                            "completedAt": "2026-04-02T08:01:00+03:00",
                            "status": "completed",
                            "cycleTimeSec": 10,
                            "startedBy": "OP-1",
                            "startedByName": "Ayse",
                        }
                    },
                    "orderSequence": ["WO-1"],
                    "source": {"file": "orders.json"},
                    "inventoryByProduct": {},
                },
            },
            "2026-04-02T08:01:00+03:00",
        )

        headers = [work_order_sheet.cell(1, idx).value for idx in range(1, work_order_sheet.max_column + 1)]
        self.assertEqual(work_order_sheet.cell(2, headers.index("order_id") + 1).value, "WO-1")
        self.assertEqual(work_order_sheet.cell(2, headers.index("ideal_cycle_ms") + 1).value, 10000)
        self.assertEqual(work_order_sheet.cell(2, headers.index("planned_duration_ms") + 1).value, 20000)
        self.assertEqual(work_order_sheet.cell(2, headers.index("runtime_ms") + 1).value, 50000)
        self.assertEqual(work_order_sheet.cell(2, headers.index("unplanned_downtime_ms") + 1).value, 10000)
        self.assertEqual(work_order_sheet.cell(2, headers.index("unplanned_downtime_sec") + 1).value, 10.0)
        self.assertEqual(work_order_sheet.cell(2, headers.index("performance_pct") + 1).value, 40.0)
        self.assertEqual(work_order_sheet.cell(2, headers.index("quality_pct") + 1).value, 100.0)
        self.assertEqual(work_order_sheet.cell(2, headers.index("availability_pct") + 1).value, 83.3)
        self.assertEqual(work_order_sheet.cell(2, headers.index("oee_pct") + 1).value, 33.3)

    def test_runtime_state_sync_writes_oee_snapshot_sheet_rows(self) -> None:
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl is not installed")
        from mes_web.excel_runtime import ExcelRuntimeSink, OEE_SNAPSHOT_COLUMNS

        workbook = Workbook()
        snapshot_sheet = workbook.active
        snapshot_sheet.title = "5_OEE_Anliklari"
        sink = ExcelRuntimeSink.__new__(ExcelRuntimeSink)
        sink._ensure_sheet_layout(snapshot_sheet, OEE_SNAPSHOT_COLUMNS)

        sink._sync_oee_snapshot_sheet(
            workbook,
            {
                "trend": [
                    {
                        "time": "2026-04-02T08:00:30+03:00",
                        "reason": "periodic_30s",
                        "summary": "Periyodik snapshot",
                        "oee": 61.7,
                        "availability": 80.0,
                        "performance": 90.0,
                        "quality": 95.0,
                        "mavi_s": 3,
                        "mavi_r": 0,
                        "mavi_h": 1,
                        "sari_s": 2,
                        "sari_r": 1,
                        "sari_h": 0,
                        "kirmizi_s": 1,
                        "kirmizi_r": 0,
                        "kirmizi_h": 0,
                    }
                ]
            },
        )

        headers = [snapshot_sheet.cell(1, idx).value for idx in range(1, snapshot_sheet.max_column + 1)]
        self.assertEqual(snapshot_sheet.cell(2, headers.index("snapshot_time") + 1).value, "2026-04-02T08:00:30+03:00")
        self.assertEqual(snapshot_sheet.cell(2, headers.index("oee") + 1).value, 61.7)
        self.assertEqual(snapshot_sheet.cell(2, headers.index("mavi_h") + 1).value, 1)
        self.assertEqual(snapshot_sheet.cell(2, headers.index("is_full_cycle_reference") + 1).value, 0)
        self.assertIn("reason=periodic_30s", str(snapshot_sheet.cell(2, headers.index("notes") + 1).value))

    def test_new_sheet_copies_reference_header_style(self) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ModuleNotFoundError:
            self.skipTest("openpyxl is not installed")
        from mes_web.excel_runtime import ExcelRuntimeSink, WORK_ORDER_COLUMNS

        workbook = Workbook()
        reference_sheet = workbook.active
        reference_sheet.title = "1_Olay_Logu"
        reference_sheet.cell(1, 1, "log_event_id")
        reference_sheet["A1"].font = Font(bold=True, color="FFFFFF")
        reference_sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")

        target_sheet = workbook.create_sheet(WORK_ORDER_SHEET_NAME)
        sink = ExcelRuntimeSink.__new__(ExcelRuntimeSink)
        sink._ensure_sheet_layout(target_sheet, WORK_ORDER_COLUMNS, workbook=workbook)

        self.assertTrue(target_sheet["A1"].font.bold)
        self.assertEqual(target_sheet["A1"].fill.fill_type, "solid")

    def test_sink_recovers_from_invalid_existing_workbook(self) -> None:
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.utils.exceptions import InvalidFileException
        except ModuleNotFoundError:
            self.skipTest("openpyxl is not installed")
        from mes_web.config import AppConfig
        from mes_web.excel_runtime import ExcelRuntimeSink

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "broken.xlsx"
            workbook_path.write_text("not-a-real-xlsx", encoding="utf-8")
            with patch.dict(
                os.environ,
                {
                    "MES_WEB_EXCEL_WORKBOOK_PATH": str(workbook_path),
                    "MES_WEB_EXCEL_TEMPLATE_PATH": str(Path(temp_dir) / "missing-template.xlsx"),
                },
                clear=False,
            ):
                config = AppConfig.from_env()
                sink = ExcelRuntimeSink(config)
                workbook = sink._open_or_create_workbook(
                    workbook_path=workbook_path,
                    workbook_factory=Workbook,
                    workbook_loader=load_workbook,
                    invalid_file_error=InvalidFileException,
                )

            self.assertIsNotNone(workbook)
            self.assertFalse(workbook_path.exists())
            archived = list(Path(temp_dir).glob("broken.xlsx.corrupt-*"))
            self.assertTrue(archived)

    def test_kiosk_event_creates_audit_log_rows(self) -> None:
        rows = self.projector.consume_kiosk_event(
            {
                "event_type": "help_requested",
                "device_id": "kiosk-1",
                "bound_station_id": "4",
                "operator_id": "1",
                "repeat_count": 2,
            },
            "2026-04-02T10:20:00Z",
        )

        self.assertEqual(rows["1_Olay_Logu"][0]["event_type_code"], "help_requested")
        self.assertEqual(rows["1_Olay_Logu"][0]["source_code"], "tablet")
        self.assertIn("repeat_count=2", rows["1_Olay_Logu"][0]["notes"])
        self.assertEqual(rows[RAW_LOG_SHEET_NAME][0]["event_type_code"], "help_requested")

    def test_fault_sheet_sync_upserts_kiosk_fault_rows(self) -> None:
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl is not installed")
        from mes_web.excel_runtime import ExcelRuntimeSink, FAULT_COLUMNS

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "3_Arizalar"
        sink = ExcelRuntimeSink.__new__(ExcelRuntimeSink)
        sink._ensure_sheet_layout(sheet, FAULT_COLUMNS, workbook=workbook)

        sink._sync_fault_sheet(
            workbook,
            {
                "shift": {"code": "SHIFT-A"},
                "activeFault": {
                    "faultId": "F-1",
                    "reasonCode": "robot_arm_jam",
                    "category": "MEKANIK",
                    "reason": "Robot Kol Sikisti",
                    "startedAt": "2026-04-02T10:00:00+03:00",
                    "source": "kiosk",
                    "deviceId": "kiosk-1",
                    "operatorId": "1",
                    "operatorCode": "OP-001",
                    "operatorName": "Test",
                    "boundStationId": "4",
                },
                "faultHistory": [],
            },
        )

        headers = [sheet.cell(1, idx).value for idx in range(1, sheet.max_column + 1)]
        self.assertEqual(sheet.cell(2, headers.index("fault_id") + 1).value, "F-1")
        self.assertEqual(sheet.cell(2, headers.index("status_code") + 1).value, "AKTIF")
        self.assertEqual(sheet.cell(2, headers.index("duration_ms") + 1).value, 0)
        self.assertEqual(sheet.cell(2, headers.index("operator_code") + 1).value, "OP-001")

    def test_maintenance_sheet_sync_writes_checklist_rows(self) -> None:
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl is not installed")
        from mes_web.excel_runtime import ExcelRuntimeSink, MAINTENANCE_COLUMNS

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = MAINTENANCE_SHEET_NAME
        sink = ExcelRuntimeSink.__new__(ExcelRuntimeSink)
        sink._ensure_sheet_layout(sheet, MAINTENANCE_COLUMNS, workbook=workbook)

        sink._sync_maintenance_sheet(
            workbook,
            {
                "maintenance": {
                    "history": [
                        {
                            "sessionId": "M-1",
                            "phase": "opening",
                            "shiftCode": "SHIFT-A",
                            "startedAt": "2026-04-02T09:00:00+03:00",
                            "endedAt": "2026-04-02T09:05:00+03:00",
                            "deviceId": "kiosk-1",
                            "deviceName": "Tablet 1",
                            "deviceRole": "operator_kiosk",
                            "boundStationId": "4",
                            "operatorId": "1",
                            "operatorCode": "OP-001",
                            "operatorName": "Test",
                            "note": "tamam",
                            "steps": [
                                {
                                    "stepCode": "opening_1",
                                    "stepLabel": "Guvenlik",
                                    "required": True,
                                    "completed": True,
                                    "completedAt": "2026-04-02T09:02:00+03:00",
                                }
                            ],
                        }
                    ]
                }
            },
        )

        headers = [sheet.cell(1, idx).value for idx in range(1, sheet.max_column + 1)]
        self.assertEqual(sheet.cell(2, headers.index("maintenance_row_key") + 1).value, "M-1:opening_1")
        self.assertEqual(sheet.cell(2, headers.index("phase_code") + 1).value, "opening")
        self.assertEqual(sheet.cell(2, headers.index("completed_flag") + 1).value, 1)
        self.assertEqual(sheet.cell(2, headers.index("duration_ms") + 1).value, 300000)
        self.assertEqual(sheet.cell(2, headers.index("note") + 1).value, "tamam")


if __name__ == "__main__":
    unittest.main()
